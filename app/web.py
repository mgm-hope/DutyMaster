from __future__ import annotations

import json
import os
import secrets
import shutil
import sqlite3
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Iterator

from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware

from .constants import (
    APP_DIR,
    DATA_DIR,
    DUTY_LABELS,
    DUTY_ORDER,
    DUTY_SECTIONS,
    EXPORTS_DIR,
    MASTER_TEMPLATE_PATH,
    ROTA_DAYS,
    ROTA_WEEKS,
    UPLOADS_DIR,
)
from .database import DB_PATH, get_connection, initialise_database
from .services import (
    assign_staff,
    active_duty_codes,
    active_duty_sections,
    auto_assign_empty_slots,
    available_staff,
    blank_duty_reason,
    build_master_style_workbook,
    candidate_rejection_reason,
    clear_inactive_teacher_break_slots,
    create_throttled_autosave,
    create_version_snapshot,
    fairness_summary,
    get_setting,
    ensure_duty_event_rows,
    get_p4_lunch_conflicts,
    preview_auto_assign,
    proposed_review,
    parse_timetable,
    reset_upload_data,
    repair_p4_lunch_conflicts,
    restore_version_snapshot,
    strict_assignment_allowed,
    save_parsed_timetable,
    workload_summary,
)


PACKAGE_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = PACKAGE_DIR / "templates"
STATIC_DIR = PACKAGE_DIR / "static"


def _ensure_bootstrap() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    conn = initialise_database(DB_PATH)
    ensure_duty_event_rows(conn)
    conn.close()


def _new_connection() -> sqlite3.Connection:
    conn = get_connection(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


@contextmanager
def _conn_context() -> Iterator[sqlite3.Connection]:
    conn = _new_connection()
    try:
        yield conn
    finally:
        conn.close()


def _visible_assignment_where(conn: sqlite3.Connection) -> tuple[str, tuple[str, ...]]:
    codes = sorted(active_duty_codes(conn))
    if not codes:
        return "1 = 0", ()
    placeholders = ", ".join("?" for _ in codes)
    return f"period IN ({placeholders})", tuple(codes)


def _prebuilt_clear_counts(conn: sqlite3.Connection) -> dict[str, int]:
    visible_where, params = _visible_assignment_where(conn)
    auto = conn.execute(
        f"""
        SELECT COUNT(*) AS count
        FROM rota_assignments
        WHERE staff_initials IS NOT NULL
          AND assignment_source = 'auto'
          AND {visible_where}
        """,
        params,
    ).fetchone()["count"]
    manual = conn.execute(
        f"""
        SELECT COUNT(*) AS count
        FROM rota_assignments
        WHERE staff_initials IS NOT NULL
          AND COALESCE(assignment_source, 'prebuilt') IN ('manual', 'prebuilt')
          AND {visible_where}
        """,
        params,
    ).fetchone()["count"]
    return {"auto": auto, "manual": manual}


STAFF_TIMETABLE_COLUMNS = [
    ("Tutor", "Tutor"),
    ("1", "Period 1"),
    ("2", "Period 2"),
    ("Break", "Break"),
    ("TeacherBreak", "Teacher Break"),
    ("3", "Period 3"),
    ("4", "Period 4 / Lunch"),
    ("5", "Period 5"),
    ("6", "Period 6"),
    ("7", "Period 7"),
]


def _staff_timetable_group_for_duty(code: str) -> str:
    if code.startswith("Tutor_"):
        return "Tutor"
    if code.startswith("P1_"):
        return "1"
    if code.startswith("P2_"):
        return "2"
    if code.startswith("Break_"):
        return "Break"
    if code.startswith("Teacher_Break_Rota_"):
        return "TeacherBreak"
    if code.startswith("P3_"):
        return "3"
    if code.startswith("P4"):
        return "4"
    if code.startswith("P5_"):
        return "5"
    if code.startswith("P6_"):
        return "6"
    if code.startswith("P7_"):
        return "7"
    if code == "Gate":
        return "Tutor"
    return "Other"


def _password_value() -> str:
    return os.getenv("DUTYMASTER_PASSWORD", "changeme123!")


def _is_logged_in(request: Request) -> bool:
    return bool(request.session.get("logged_in"))


def _require_login(request: Request) -> RedirectResponse | None:
    if _is_logged_in(request):
        return None
    return RedirectResponse("/login", status_code=303)


def _flash(request: Request, message: str) -> None:
    request.session["flash"] = message


def _get_flash(request: Request) -> str:
    return request.session.pop("flash", "")


def _base_context(request: Request, title: str, **extra) -> dict:
    return {
        "request": request,
        "title": title,
        "flash": _get_flash(request),
        "is_logged_in": _is_logged_in(request),
        "days": ROTA_DAYS,
        "weeks": ROTA_WEEKS,
        **extra,
    }


def _save_upload(upload: UploadFile, prefix: str) -> Path:
    safe_name = Path(upload.filename or f"{prefix}.xlsx").name
    target = UPLOADS_DIR / f"{prefix}_{secrets.token_hex(4)}_{safe_name}"
    with target.open("wb") as handle:
        shutil.copyfileobj(upload.file, handle)
    return target


_ensure_bootstrap()

app = FastAPI(title="DutyMaster Online")
app.add_middleware(SessionMiddleware, secret_key=os.getenv("DUTYMASTER_SECRET_KEY", "replace-me-on-railway"))
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    if not _is_logged_in(request):
        return RedirectResponse("/login", status_code=303)
    return RedirectResponse("/dashboard", status_code=303)


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", _base_context(request, "Login"))


@app.post("/login")
async def login(request: Request, password: str = Form(...)):
    if secrets.compare_digest(password, _password_value()):
        request.session["logged_in"] = True
        _flash(request, "Signed in.")
        return RedirectResponse("/dashboard", status_code=303)
    _flash(request, "Password not recognised.")
    return RedirectResponse("/login", status_code=303)


@app.post("/logout")
def logout(request: Request):
    if _is_logged_in(request):
        with _conn_context() as conn:
            create_version_snapshot(conn, f"Logout save {datetime.now().strftime('%d %b %H:%M')}", "Saved on logout")
    request.session.clear()
    return RedirectResponse("/login", status_code=303)


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        meta = conn.execute("SELECT * FROM timetable_meta ORDER BY id DESC LIMIT 1").fetchone()
        teacher_count = conn.execute("SELECT COUNT(*) AS count FROM teachers").fetchone()["count"]
        additional_count = conn.execute("SELECT COUNT(*) AS count FROM additional_staff WHERE COALESCE(is_archived, 0)=0").fetchone()["count"]
        assigned_count = conn.execute("SELECT COUNT(*) AS count FROM rota_assignments WHERE staff_initials IS NOT NULL").fetchone()["count"]
        conflicts = get_p4_lunch_conflicts(conn)
        workload = workload_summary(conn)[:12]
        fairness = fairness_summary(workload_summary(conn))
    return templates.TemplateResponse(
        "dashboard.html",
        _base_context(
            request,
            "Dashboard",
            meta=meta,
            teacher_count=teacher_count,
            additional_count=additional_count,
            assigned_count=assigned_count,
            conflicts=conflicts,
            workload=workload,
            fairness=fairness,
            db_path=str(DB_PATH),
        ),
    )


@app.get("/uploads", response_class=HTMLResponse)
def uploads(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        meta = conn.execute("SELECT * FROM timetable_meta ORDER BY id DESC LIMIT 1").fetchone()
    return templates.TemplateResponse(
        "uploads.html",
        _base_context(request, "Uploads", meta=meta, template_path=str(MASTER_TEMPLATE_PATH)),
    )


@app.post("/uploads/timetable")
async def upload_timetable(request: Request, timetable: UploadFile = File(...)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    target = _save_upload(timetable, "timetable")
    parsed = parse_timetable(target)
    with _conn_context() as conn:
        create_version_snapshot(conn, f"Before timetable upload {datetime.now().strftime('%d %b %H:%M')}", "Safety copy before timetable upload")
        save_parsed_timetable(conn, parsed)
    _flash(request, f"Uploaded timetable: {len(parsed['teachers'])} teachers parsed. Teaching Loads is ready to review.")
    return RedirectResponse("/teaching-loads", status_code=303)


@app.post("/uploads/staff-list")
async def upload_staff_list(request: Request, staff_list: UploadFile = File(...)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    target = _save_upload(staff_list, "staff_list")
    from openpyxl import load_workbook
    import csv

    rows = []
    if target.suffix.lower() == ".csv":
        with target.open(newline="", encoding="utf-8-sig") as handle:
            rows = list(csv.DictReader(handle))
    else:
        wb = load_workbook(target)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        for values in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, values)))
    count = 0
    with _conn_context() as conn:
        create_version_snapshot(conn, f"Before staff list upload {datetime.now().strftime('%d %b %H:%M')}", "Safety copy before staff list upload")
        for row in rows:
            initials = str(row.get("Initials") or row.get("initials") or "").strip().upper()
            full_name = str(row.get("Full Name") or row.get("full_name") or row.get("Name") or "").strip()
            if initials and full_name:
                conn.execute(
                    "INSERT OR REPLACE INTO staff_names(initials, full_name, last_updated) VALUES (?, ?, ?)",
                    (initials, full_name, datetime.now().isoformat()),
                )
                conn.execute("UPDATE teachers SET full_name = ? WHERE initials = ?", (full_name, initials))
                count += 1
        conn.commit()
    _flash(request, f"Imported {count} staff names.")
    return RedirectResponse("/uploads", status_code=303)


@app.post("/uploads/master-template")
async def upload_master_template(request: Request, master_template: UploadFile = File(...)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    MASTER_TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with MASTER_TEMPLATE_PATH.open("wb") as handle:
        shutil.copyfileobj(master_template.file, handle)
    _flash(request, "Master duty template uploaded for exports.")
    return RedirectResponse("/uploads", status_code=303)


@app.post("/uploads/reset")
def reset_upload(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_version_snapshot(conn, f"Before upload reset {datetime.now().strftime('%d %b %H:%M')}", "Safety copy before upload reset")
        reset_upload_data(conn)
    _flash(request, "Uploaded timetable data reset. Additional staffing was kept.")
    return RedirectResponse("/uploads", status_code=303)


@app.get("/teaching-loads", response_class=HTMLResponse)
def teaching_loads(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    try:
        with _conn_context() as conn:
            teachers = conn.execute(
                """
                SELECT t.initials,
                       COALESCE(s.full_name, t.full_name, t.initials) AS full_name,
                       COALESCE(t.total_lessons, 0) AS total_lessons,
                       COALESCE(t.protected_periods, 6) AS protected_periods,
                       COALESCE(t.classification, 'Teacher') AS classification,
                       COALESCE(t.is_part_time, 0) AS is_part_time,
                       COALESCE(t.days_in_school, '1111111111') AS days_in_school,
                       COALESCE(t.subject, '') AS subject,
                       t.max_lunch_duties,
                       COALESCE(t.min_duties, 0) AS min_duties,
                       COALESCE(t.can_first_duty, 0) AS can_first_duty,
                       COALESCE(t.exclude_from_algorithm, 0) AS exclude_from_algorithm
                FROM teachers t
                LEFT JOIN staff_names s ON t.initials = s.initials
                ORDER BY t.initials
                """
            ).fetchall()
        return templates.TemplateResponse("teaching_loads.html", _base_context(request, "Teaching Loads", teachers=teachers))
    except Exception as exc:
        return templates.TemplateResponse(
            "error.html",
            _base_context(
                request,
                "Teaching Loads Error",
                heading="Teaching Loads could not open",
                message=str(exc),
                next_step="Try re-uploading the timetable. If this persists, copy this message from the Railway app.",
            ),
            status_code=500,
        )


@app.post("/teaching-loads/update")
async def update_teacher(
    request: Request,
    initials: str = Form(...),
    protected_periods: int = Form(...),
    classification: str = Form(...),
    days_in_school: str = Form(...),
    subject: str = Form(""),
    max_lunch_duties: str = Form(""),
    min_duties: int = Form(0),
    can_first_duty: int = Form(0),
    exclude_from_algorithm: int = Form(0),
):
    redirect = _require_login(request)
    if redirect:
        return redirect
    days = "".join("1" if char == "1" else "0" for char in days_in_school)[:10].ljust(10, "1")
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Teaching load edit autosave")
        total = conn.execute("SELECT total_lessons FROM teachers WHERE initials = ?", (initials,)).fetchone()["total_lessons"]
        days_out = days.count("0")
        max_load = 65 - (6.5 * days_out)
        lunch_limit = int(max_lunch_duties) if str(max_lunch_duties).strip() else None
        conn.execute(
            """
            UPDATE teachers
            SET protected_periods = ?, classification = ?, days_in_school = ?,
                is_part_time = ?, non_contact = ?, subject = ?, max_lunch_duties = ?,
                min_duties = ?, can_first_duty = ?, exclude_from_algorithm = ?, last_updated = ?
            WHERE initials = ?
            """,
            (
                protected_periods,
                classification,
                days,
                1 if "0" in days else 0,
                max(0, max_load - float(total or 0)),
                subject.strip(),
                lunch_limit,
                max(0, min_duties),
                1 if can_first_duty else 0,
                1 if exclude_from_algorithm else 0,
                datetime.now().isoformat(),
                initials,
            ),
        )
        conn.commit()
    _flash(request, f"Updated {initials}.")
    return RedirectResponse("/teaching-loads", status_code=303)


@app.get("/staff-load-summary", response_class=HTMLResponse)
def staff_load_summary(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    selected_staff = (request.query_params.get("staff") or "").strip().upper()
    teaching_periods = {"1", "2", "3", "4", "5", "6"}
    lesson_duty_prefixes = ("P1_", "P2_", "P3_", "P4_", "P4A_", "P4B_", "P4C_", "P5_", "P6_")
    with _conn_context() as conn:
        teachers = conn.execute(
            """
            SELECT t.initials,
                   COALESCE(s.full_name, t.full_name, t.initials) AS full_name,
                   COALESCE(t.classification, 'Teacher') AS classification,
                   COALESCE(t.protected_periods, 6) AS protected_periods,
                   COALESCE(t.days_in_school, '1111111111') AS days_in_school,
                   COALESCE(t.subject, '') AS subject
            FROM teachers t
            LEFT JOIN staff_names s ON t.initials = s.initials
            ORDER BY t.initials
            """
        ).fetchall()
        rows = []
        totals = {
            "teaching": 0,
            "non_contact": 0,
            "protected": 0,
            "p1_6_duties": 0,
            "break_duties": 0,
            "first_duties": 0,
            "lunch_duties": 0,
            "remaining": 0,
        }
        teacher_options = [
            {"initials": row["initials"], "label": f"{row['initials']} - {row['full_name']} ({row['classification']})"}
            for row in teachers
        ]
        for teacher in teachers:
            initials = teacher["initials"]
            days = (teacher["days_in_school"] or "1111111111").ljust(10, "1")[:10]
            possible_p1_6 = days.count("1") * 6
            teaching_p1_6 = conn.execute(
                """
                SELECT COUNT(*) AS count
                FROM teacher_periods
                WHERE teacher_initials = ?
                  AND period IN ('1', '2', '3', '4', '5', '6')
                """,
                (initials,),
            ).fetchone()["count"]
            duty_rows = conn.execute(
                """
                SELECT period
                FROM rota_assignments
                WHERE staff_initials = ?
                """,
                (initials,),
            ).fetchall()
            duty_codes = [row["period"] for row in duty_rows]
            p1_6_duties = sum(1 for code in duty_codes if code.startswith(lesson_duty_prefixes))
            break_duties = sum(1 for code in duty_codes if code.startswith("Break_") or code.startswith("Teacher_Break_Rota_"))
            first_duties = sum(1 for code in duty_codes if "First_Duty" in code)
            lunch_duties = sum(
                1
                for code in duty_codes
                if code.startswith("P4_Lunch_") or code in {"P4A_Lunch_Duty", "P4B_Lunch_Duty", "P4C_Lunch_Duty"}
            )
            total_duties = len(duty_codes)
            non_contact_p1_6 = max(0, possible_p1_6 - int(teaching_p1_6 or 0))
            protected = int(teacher["protected_periods"] or 0)
            remaining = non_contact_p1_6 - protected - p1_6_duties
            row = {
                "initials": initials,
                "full_name": teacher["full_name"],
                "classification": teacher["classification"],
                "subject": teacher["subject"] or "",
                "possible_p1_6": possible_p1_6,
                "teaching_p1_6": int(teaching_p1_6 or 0),
                "non_contact_p1_6": non_contact_p1_6,
                "protected": protected,
                "p1_6_duties": p1_6_duties,
                "break_duties": break_duties,
                "first_duties": first_duties,
                "lunch_duties": lunch_duties,
                "total_duties": total_duties,
                "remaining": remaining,
            }
            rows.append(row)
            totals["teaching"] += row["teaching_p1_6"]
            totals["non_contact"] += row["non_contact_p1_6"]
            totals["protected"] += row["protected"]
            totals["p1_6_duties"] += row["p1_6_duties"]
            totals["break_duties"] += row["break_duties"]
            totals["first_duties"] += row["first_duties"]
            totals["lunch_duties"] += row["lunch_duties"]
            totals["remaining"] += row["remaining"]
        selected_teacher = None
        replacement_options = []
        if selected_staff:
            selected_teacher = conn.execute(
                """
                SELECT t.initials,
                       COALESCE(s.full_name, t.full_name, t.initials) AS full_name,
                       COALESCE(t.classification, 'Teacher') AS classification
                FROM teachers t
                LEFT JOIN staff_names s ON t.initials = s.initials
                WHERE t.initials = ?
                """,
                (selected_staff,),
            ).fetchone()
        ignored = set(request.session.get("load_summary_ignored", []))
        if selected_teacher:
            for duty in conn.execute(
                """
                SELECT week, day, period, staff_initials, staff_type, COALESCE(assignment_source, '') AS assignment_source
                FROM rota_assignments
                WHERE staff_initials IS NOT NULL
                  AND staff_type = 'SLT'
                ORDER BY week, id
                """
            ).fetchall():
                key = f"{selected_staff}|{duty['week']}|{duty['day']}|{duty['period']}"
                if key in ignored:
                    continue
                if not strict_assignment_allowed(
                    conn,
                    selected_teacher["initials"],
                    selected_teacher["classification"],
                    duty["week"],
                    duty["day"],
                    duty["period"],
                    "teacher",
                ):
                    continue
                replacement_options.append(
                    {
                        "key": key,
                        "week": duty["week"],
                        "day": duty["day"],
                        "period": duty["period"],
                        "label": DUTY_LABELS.get(duty["period"], duty["period"]),
                        "current": duty["staff_initials"],
                        "source": duty["assignment_source"],
                    }
                )
    return templates.TemplateResponse(
        "staff_load_summary.html",
        _base_context(
            request,
            "Staff Load Summary",
            rows=rows,
            totals=totals,
            teacher_options=teacher_options,
            selected_staff=selected_staff,
            selected_teacher=selected_teacher,
            replacement_options=replacement_options,
        ),
    )


@app.post("/staff-load-summary/assign")
async def staff_load_summary_assign(
    request: Request,
    initials: str = Form(...),
    week: int = Form(...),
    day: str = Form(...),
    period: str = Form(...),
):
    redirect = _require_login(request)
    if redirect:
        return redirect
    initials = initials.strip().upper()
    with _conn_context() as conn:
        teacher = conn.execute("SELECT classification FROM teachers WHERE initials = ?", (initials,)).fetchone()
        if not teacher:
            _flash(request, f"{initials} was not found.")
            return RedirectResponse(f"/staff-load-summary?staff={initials}", status_code=303)
        current = conn.execute(
            "SELECT staff_initials, staff_type FROM rota_assignments WHERE week = ? AND day = ? AND period = ?",
            (week, day, period),
        ).fetchone()
        if not current or current["staff_type"] != "SLT":
            _flash(request, "That duty is no longer held by SLT.")
            return RedirectResponse(f"/staff-load-summary?staff={initials}", status_code=303)
        result = assign_staff(conn, week, day, period, initials, teacher["classification"], assignment_source="manual_adjustment")
        if result < 0:
            _flash(request, f"{initials} could not be assigned without breaking an active rule.")
        else:
            _flash(request, f"{initials} replaced {current['staff_initials']} on {DUTY_LABELS.get(period, period)}.")
    return RedirectResponse(f"/staff-load-summary?staff={initials}", status_code=303)


@app.post("/staff-load-summary/ignore")
async def staff_load_summary_ignore(
    request: Request,
    initials: str = Form(...),
    week: int = Form(...),
    day: str = Form(...),
    period: str = Form(...),
):
    redirect = _require_login(request)
    if redirect:
        return redirect
    initials = initials.strip().upper()
    key = f"{initials}|{week}|{day}|{period}"
    ignored = set(request.session.get("load_summary_ignored", []))
    ignored.add(key)
    request.session["load_summary_ignored"] = sorted(ignored)
    _flash(request, f"Ignored this suggestion for {initials}.")
    return RedirectResponse(f"/staff-load-summary?staff={initials}", status_code=303)


@app.get("/additional-staff", response_class=HTMLResponse)
def additional_staff(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        staff = conn.execute(
            """
            SELECT id, category, initials, full_name, is_full_time,
                   COALESCE(days_in_school, '1111111111') AS days_in_school,
                   COALESCE(availability, '[]') AS availability,
                   is_archived, COALESCE(status, 'Active') AS status,
                   COALESCE(min_duties, 0) AS min_duties,
                   max_duties
            FROM additional_staff
            ORDER BY is_archived, category, initials
            """
        ).fetchall()
    period_options = [
        ("Tutor", "Tutor"),
        ("1", "Period 1"),
        ("2", "Period 2"),
        ("Break", "Break"),
        ("TeacherBreak", "Teaching Staff Break Rota"),
        ("3", "Period 3"),
        ("4", "Period 4 / Lunch"),
        ("5", "Period 5"),
        ("6", "Period 6"),
        ("7", "Period 7"),
    ]
    staff_rows = []
    for row in staff:
        item = dict(row)
        try:
            item["period_availability"] = set(json.loads(item.get("availability") or "[]"))
        except (TypeError, json.JSONDecodeError):
            item["period_availability"] = set()
        staff_rows.append(item)
    return templates.TemplateResponse(
        "additional_staff.html",
        _base_context(request, "Additional Staffing", staff=staff_rows, period_options=period_options),
    )


@app.post("/additional-staff/add")
async def add_additional_staff(
    request: Request,
    initials: str = Form(...),
    full_name: str = Form(...),
    category: str = Form(...),
):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Additional staffing edit autosave")
        try:
            conn.execute(
                """
                INSERT INTO additional_staff(category, initials, full_name, is_full_time, days_in_school, availability, min_duties, max_duties, last_updated)
                VALUES (?, ?, ?, 1, '1111111111', '[]', 0, NULL, ?)
                """,
                (category, initials.strip().upper(), full_name.strip(), datetime.now().isoformat()),
            )
            conn.commit()
            _flash(request, f"Added {initials.upper()}.")
        except sqlite3.IntegrityError:
            _flash(request, f"{initials.upper()} already exists.")
    return RedirectResponse("/additional-staff", status_code=303)


@app.post("/additional-staff/status")
async def update_additional_status(request: Request, staff_id: int = Form(...), status: str = Form(...)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Additional staffing edit autosave")
        conn.execute("UPDATE additional_staff SET status = ?, last_updated = ? WHERE id = ?", (status, datetime.now().isoformat(), staff_id))
        conn.commit()
    return RedirectResponse("/additional-staff", status_code=303)


@app.post("/additional-staff/availability")
async def update_additional_availability(
    request: Request,
    staff_id: int = Form(...),
    days_in_school: str = Form(...),
    periods_available: list[str] = Form([]),
    min_duties: int = Form(0),
    max_duties: str = Form(""),
):
    redirect = _require_login(request)
    if redirect:
        return redirect
    days = "".join("1" if char == "1" else "0" for char in days_in_school)[:10].ljust(10, "1")
    max_limit = int(max_duties) if str(max_duties).strip() else None
    allowed_periods = {"Tutor", "1", "2", "Break", "TeacherBreak", "3", "4", "5", "6", "7"}
    period_list = [period for period in periods_available if period in allowed_periods]
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Additional staffing edit autosave")
        conn.execute(
            """
            UPDATE additional_staff
            SET days_in_school = ?, availability = ?, is_full_time = ?, min_duties = ?, max_duties = ?, last_updated = ?
            WHERE id = ?
            """,
            (
                days,
                json.dumps(period_list),
                0 if "0" in days else 1,
                max(0, min_duties),
                max_limit,
                datetime.now().isoformat(),
                staff_id,
            ),
        )
        conn.commit()
    _flash(request, "Additional staff availability updated.")
    return RedirectResponse("/additional-staff", status_code=303)


@app.post("/additional-staff/archive")
async def archive_additional(request: Request, staff_id: int = Form(...), archive: int = Form(...)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_version_snapshot(conn, f"Before archive change {datetime.now().strftime('%d %b %H:%M')}", "Safety copy before additional staff archive/restore")
        conn.execute("UPDATE additional_staff SET is_archived = ?, last_updated = ? WHERE id = ?", (archive, datetime.now().isoformat(), staff_id))
        conn.commit()
    return RedirectResponse("/additional-staff", status_code=303)


@app.get("/prebuilt", response_class=HTMLResponse)
def prebuilt(request: Request, week: int = 1, day: str = "Mon", period: str | None = None):
    redirect = _require_login(request)
    if redirect:
        return redirect
    if week not in ROTA_WEEKS:
        week = 1
    if day not in ROTA_DAYS:
        day = "Mon"
    valid_periods = set(DUTY_LABELS)
    if period not in valid_periods:
        period = None
    selected_section = ""
    if period:
        for section_name, events in DUTY_SECTIONS:
            if any(code == period for code, _ in events):
                selected_section = section_name
                break
    day_names = {"Mon": "Monday", "Tue": "Tuesday", "Wed": "Wednesday", "Thu": "Thursday", "Fri": "Friday"}
    with _conn_context() as conn:
        ensure_duty_event_rows(conn)
        assignments = {
            f"{row['week']}:{row['day']}:{row['period']}": row["staff_initials"] or ""
            for row in conn.execute("SELECT week, day, period, staff_initials FROM rota_assignments").fetchall()
        }
        candidates = available_staff(conn, week, day, period) if period else []
        conflicts = get_p4_lunch_conflicts(conn)
        preview = preview_auto_assign(conn) if request.query_params.get("preview") == "1" else None
        duty_sections = active_duty_sections(conn)
        clear_counts = _prebuilt_clear_counts(conn)
    return templates.TemplateResponse(
        "prebuilt.html",
        _base_context(
            request,
            "Pre-Built Duty Events",
            duty_sections=duty_sections,
            duty_labels=DUTY_LABELS,
            assignments=assignments,
            selected={
                "week": week,
                "day": day,
                "day_name": day_names.get(day, day),
                "period": period,
                "section": selected_section,
            },
            candidates=candidates,
            conflicts=conflicts,
            preview=preview,
            clear_counts=clear_counts,
        ),
    )


@app.post("/prebuilt/assign")
async def prebuilt_assign(
    request: Request,
    week: int = Form(...),
    day: str = Form(...),
    period: str = Form(...),
    staff_value: str = Form(...),
    scope: str = Form("single"),
):
    redirect = _require_login(request)
    if redirect:
        return redirect
    initials, role = staff_value.split("|", 1)
    targets = [(week, day)]
    if scope == "week":
        targets = [(week, target_day) for target_day in ROTA_DAYS]
    elif scope == "both":
        targets = [(target_week, target_day) for target_week in ROTA_WEEKS for target_day in ROTA_DAYS]
    assigned = 0
    skipped = 0
    skip_details = []
    cleared = 0
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Pre-built assignment autosave")
        for target_week, target_day in targets:
            possible = available_staff(conn, target_week, target_day, period)
            if not any(item["initials"] == initials and item["role"] == role for item in possible):
                skipped += 1
                reason = candidate_rejection_reason(conn, initials, role, target_week, target_day, period)
                skip_details.append(f"W{target_week} {target_day}: {reason or 'not available for this duty'}")
                continue
            result = assign_staff(conn, target_week, target_day, period, initials, role, assignment_source="prebuilt")
            if result < 0:
                skipped += 1
                reason = candidate_rejection_reason(conn, initials, role, target_week, target_day, period)
                skip_details.append(f"W{target_week} {target_day}: {reason or 'assignment blocked by active rules'}")
                continue
            cleared += result
            assigned += 1
    clear_note = f" Cleared {cleared} Period 4 clash(es)." if cleared else ""
    skip_note = ""
    if skipped:
        shown = "; ".join(skip_details[:6])
        more = f"; plus {len(skip_details) - 6} more" if len(skip_details) > 6 else ""
        skip_note = f" Skipped {skipped} session(s): {shown}{more}."
    _flash(request, f"Assigned {initials} to {assigned} session(s).{skip_note}{clear_note}")
    return RedirectResponse(f"/prebuilt?week={week}&day={day}&period={period}", status_code=303)


@app.post("/prebuilt/clear")
async def prebuilt_clear(request: Request, week: int = Form(...), day: str = Form(...), period: str = Form(...)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Pre-built assignment autosave")
        conn.execute(
            "UPDATE rota_assignments SET staff_initials = NULL, staff_type = NULL, assignment_source = NULL, last_updated = ? WHERE week = ? AND day = ? AND period = ?",
            (datetime.now().isoformat(), week, day, period),
        )
        conn.commit()
    return RedirectResponse(f"/prebuilt?week={week}&day={day}&period={period}", status_code=303)


@app.post("/prebuilt/clear-all")
async def prebuilt_clear_all(request: Request, confirm: str = Form("")):
    redirect = _require_login(request)
    if redirect:
        return redirect
    if confirm.strip().upper() != "CLEAR":
        _flash(request, "Type CLEAR to confirm clearing all assignments.")
        return RedirectResponse("/prebuilt", status_code=303)
    with _conn_context() as conn:
        create_version_snapshot(conn, f"Before clear all {datetime.now().strftime('%d %b %H:%M')}", "Safety copy before clearing all assignments")
        conn.execute("UPDATE rota_assignments SET staff_initials = NULL, staff_type = NULL, assignment_source = NULL, last_updated = ?", (datetime.now().isoformat(),))
        conn.commit()
    _flash(request, "All duty assignments cleared.")
    return RedirectResponse("/prebuilt", status_code=303)


@app.post("/prebuilt/clear-auto")
async def prebuilt_clear_auto(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_version_snapshot(conn, f"Before clear auto {datetime.now().strftime('%d %b %H:%M')}", "Safety copy before clearing auto-assigned duties")
        visible_where, params = _visible_assignment_where(conn)
        cursor = conn.execute(
            f"""
            UPDATE rota_assignments
            SET staff_initials = NULL, staff_type = NULL, assignment_source = NULL, last_updated = ?
            WHERE assignment_source = 'auto'
              AND staff_initials IS NOT NULL
              AND {visible_where}
            """,
            (datetime.now().isoformat(), *params),
        )
        conn.commit()
    _flash(request, f"Cleared {cursor.rowcount} auto-assigned duty assignment(s).")
    return RedirectResponse("/prebuilt", status_code=303)


@app.post("/prebuilt/clear-manual")
async def prebuilt_clear_manual(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_version_snapshot(conn, f"Before clear manual {datetime.now().strftime('%d %b %H:%M')}", "Safety copy before clearing manual assignments")
        visible_where, params = _visible_assignment_where(conn)
        cursor = conn.execute(
            f"""
            UPDATE rota_assignments
            SET staff_initials = NULL, staff_type = NULL, assignment_source = NULL, last_updated = ?
            WHERE COALESCE(assignment_source, 'prebuilt') IN ('manual', 'prebuilt')
              AND staff_initials IS NOT NULL
              AND {visible_where}
            """,
            (datetime.now().isoformat(), *params),
        )
        conn.commit()
    _flash(request, f"Cleared {cursor.rowcount} manual duty assignment(s).")
    return RedirectResponse("/prebuilt", status_code=303)


@app.post("/prebuilt/repair-p4")
async def prebuilt_repair_p4(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_version_snapshot(conn, f"Before Period 4 repair {datetime.now().strftime('%d %b %H:%M')}", "Safety copy before Period 4 repair")
        cleared = repair_p4_lunch_conflicts(conn)
    _flash(request, f"Repaired Period 4 lunch conflicts. Cleared {cleared} 4A/4B/4C assignment(s).")
    return RedirectResponse("/prebuilt", status_code=303)


@app.post("/prebuilt/auto-assign")
async def prebuilt_auto_assign(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_version_snapshot(conn, f"Before auto-assign {datetime.now().strftime('%d %b %H:%M')}", "Safety copy before auto-assign")
        result = auto_assign_empty_slots(conn)
    _flash(
        request,
        f"Auto-assigned {result['assigned']} slot(s). {result['issues']} need manual review. "
        f"Repaired {result['repaired']} Period 4 conflict(s). "
        f"Moved {result.get('replaced_first_duty_pastoral', 0)} First Duty slot(s) from SLT to Pastoral. "
        f"Made {result.get('lunch_swaps', 0)} lunch fairness swap(s).",
    )
    return RedirectResponse("/prebuilt", status_code=303)


@app.get("/prebuilt/preview", response_class=HTMLResponse)
def prebuilt_preview(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    return RedirectResponse("/prebuilt?preview=1", status_code=303)


@app.get("/rules", response_class=HTMLResponse)
def rules_page(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        rules = conn.execute(
            """
            SELECT id, name, description, active
            FROM rules
            ORDER BY
                CASE
                    WHEN name = 'Room 90 Manual Fill Only' THEN 0
                    ELSE 1
                END,
                id
            """
        ).fetchall()
        max_duties = get_setting(conn, "max_duties_per_week", "4")
        max_duties_day = get_setting(conn, "max_duties_per_day", "2")
        teacher_break_slots = get_setting(conn, "teacher_break_rota_slots", "6")
        p7_mode = get_setting(conn, "p7_mode", get_setting(conn, "p7_detention_mode", "ignore"))
        exclusions = conn.execute(
            """
            SELECT id, staff_initials, week, day, reason, active, created_at
            FROM staff_exclusions
            ORDER BY active DESC, id DESC
            LIMIT 80
            """
        ).fetchall()
        custom_rules = conn.execute(
            """
            SELECT id, name, active, duty_scope, staff_scope, condition_type, condition_value, priority, notes
            FROM custom_rules
            WHERE COALESCE(is_archived, 0) = 0
            ORDER BY active DESC, id DESC
            """
        ).fetchall()
        teacher_staff = conn.execute(
            """
            SELECT initials, COALESCE(full_name, initials) AS name, classification AS role
            FROM teachers
            ORDER BY initials
            """
        ).fetchall()
        additional_staff = conn.execute(
            """
            SELECT initials, full_name AS name, category AS role
            FROM additional_staff
            WHERE COALESCE(is_archived, 0) = 0
            ORDER BY category, initials
            """
        ).fetchall()
    staff_options = [
        {"value": f"Staff:{row['initials']}", "label": f"{row['initials']} - {row['name']} ({row['role']})"}
        for row in [*teacher_staff, *additional_staff]
    ]
    duty_scopes = [
        "Any", "Gate Duty", "Tutor Time", "Period 1", "Period 2", "Break", "Period 3",
        "Teaching Staff Break Rota", "Period 4 Lunch", "Period 4A", "Period 4B", "Period 4C", "Period 5", "Period 6",
        "Period 7", "Isolation Duties", "Lunch and Detention", "Heavy Duties",
    ]
    duty_options = [{"value": code, "label": f"{section} - {label}"} for section, events in DUTY_SECTIONS for code, label in events]
    staff_scopes = ["Any", "Teacher", "HOF", "SLT", "Pastoral", "Admin", "ESLT", "Chaplaincy"]
    condition_types = [
        ("exclude", "Exclude this staff/group from this duty scope"),
        ("min_available_periods", "Require at least this many available periods"),
        ("max_duties_per_week", "Limit this staff/group to this many duties per week"),
        ("no_heavy_same_day", "Do not give a second heavy duty on the same day"),
    ]
    return templates.TemplateResponse(
        "rules.html",
        _base_context(
            request,
            "Rules",
            rules=rules,
            max_duties=max_duties,
            max_duties_day=max_duties_day,
            teacher_break_slots=teacher_break_slots,
            p7_mode=p7_mode,
            exclusions=exclusions,
            custom_rules=custom_rules,
            duty_scopes=duty_scopes,
            duty_options=duty_options,
            staff_scopes=staff_scopes,
            staff_options=staff_options,
            condition_types=condition_types,
        ),
    )


@app.post("/rules/toggle")
async def rules_toggle(request: Request, rule_id: int = Form(...), active: int = Form(0)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Rules edit autosave")
        conn.execute(
            "UPDATE rules SET active = ?, last_updated = ? WHERE id = ?",
            (1 if active else 0, datetime.now().isoformat(), rule_id),
        )
        conn.commit()
    return RedirectResponse("/rules", status_code=303)


@app.post("/rules/settings")
async def rules_settings(
    request: Request,
    max_duties_per_week: int = Form(...),
    max_duties_per_day: int = Form(2),
    teacher_break_rota_slots: int = Form(6),
    p7_mode: str = Form("ignore"),
):
    redirect = _require_login(request)
    if redirect:
        return redirect
    value = str(max(0, min(30, max_duties_per_week)))
    day_value = str(max(0, min(10, max_duties_per_day)))
    break_slots = str(max(0, min(10, teacher_break_rota_slots)))
    period7_mode = p7_mode if p7_mode in {"ignore", "slt", "pastoral"} else "ignore"
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Rules edit autosave")
        conn.execute(
            """
            INSERT INTO app_settings(key, value, last_updated)
            VALUES ('max_duties_per_week', ?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value, last_updated = excluded.last_updated
            """,
            (value, datetime.now().isoformat()),
        )
        conn.execute(
            """
            INSERT INTO app_settings(key, value, last_updated)
            VALUES ('teacher_break_rota_slots', ?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value, last_updated = excluded.last_updated
            """,
            (break_slots, datetime.now().isoformat()),
        )
        conn.execute(
            """
            INSERT INTO app_settings(key, value, last_updated)
            VALUES ('max_duties_per_day', ?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value, last_updated = excluded.last_updated
            """,
            (day_value, datetime.now().isoformat()),
        )
        conn.execute(
            """
            INSERT INTO app_settings(key, value, last_updated)
            VALUES ('p7_mode', ?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value, last_updated = excluded.last_updated
            """,
            (period7_mode, datetime.now().isoformat()),
        )
        cleared = clear_inactive_teacher_break_slots(conn)
        conn.commit()
    clear_note = f" Cleared {cleared} inactive teacher break rota assignment(s)." if cleared else ""
    _flash(request, f"Maximum duties set to {value} per week and {day_value} per day. Teaching staff break rota slots set to {break_slots}. Period 7 mode set to {period7_mode}.{clear_note}")
    return RedirectResponse("/rules", status_code=303)


@app.post("/rules/exclusions/add")
async def staff_exclusion_add(
    request: Request,
    staff_value: str = Form(...),
    week: int = Form(...),
    day: str = Form(...),
    reason: str = Form(""),
):
    redirect = _require_login(request)
    if redirect:
        return redirect
    initials = staff_value.split("|", 1)[0]
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Rules edit autosave")
        conn.execute(
            """
            INSERT INTO staff_exclusions(staff_initials, week, day, reason, active, last_updated)
            VALUES (?, ?, ?, ?, 1, ?)
            """,
            (initials, week, day, reason.strip(), datetime.now().isoformat()),
        )
        conn.commit()
    _flash(request, f"Added one-off exclusion for {initials}.")
    return RedirectResponse("/rules", status_code=303)


@app.post("/rules/exclusions/toggle")
async def staff_exclusion_toggle(request: Request, exclusion_id: int = Form(...), active: int = Form(0)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Rules edit autosave")
        conn.execute(
            "UPDATE staff_exclusions SET active = ?, last_updated = ? WHERE id = ?",
            (1 if active else 0, datetime.now().isoformat(), exclusion_id),
        )
        conn.commit()
    return RedirectResponse("/rules", status_code=303)


@app.post("/rules/custom/add")
async def custom_rule_add(
    request: Request,
    name: str = Form(...),
    duty_scope: str = Form("Any"),
    duty_specific: str = Form(""),
    staff_scope: str = Form("Any"),
    staff_specific: str = Form(""),
    condition_type: str = Form(...),
    condition_value: str = Form(""),
    priority: str = Form("Hard"),
    notes: str = Form(""),
):
    redirect = _require_login(request)
    if redirect:
        return redirect
    final_duty_scope = duty_specific or duty_scope
    final_staff_scope = staff_specific or staff_scope
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Rules edit autosave")
        conn.execute(
            """
            INSERT INTO custom_rules(name, duty_scope, staff_scope, condition_type, condition_value, priority, notes, last_updated)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                name.strip() or "Custom rule",
                final_duty_scope,
                final_staff_scope,
                condition_type,
                condition_value.strip(),
                priority,
                notes.strip(),
                datetime.now().isoformat(),
            ),
        )
        conn.commit()
    _flash(request, "Custom rule added.")
    return RedirectResponse("/rules", status_code=303)


@app.post("/rules/custom/toggle")
async def custom_rule_toggle(request: Request, rule_id: int = Form(...), active: int = Form(0)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Rules edit autosave")
        conn.execute(
            "UPDATE custom_rules SET active = ?, last_updated = ? WHERE id = ?",
            (1 if active else 0, datetime.now().isoformat(), rule_id),
        )
        conn.commit()
    return RedirectResponse("/rules", status_code=303)


@app.post("/rules/custom/archive")
async def custom_rule_archive(request: Request, rule_id: int = Form(...)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Rules edit autosave")
        conn.execute(
            "UPDATE custom_rules SET is_archived = 1, active = 0, last_updated = ? WHERE id = ?",
            (datetime.now().isoformat(), rule_id),
        )
        conn.commit()
    _flash(request, "Custom rule removed.")
    return RedirectResponse("/rules", status_code=303)


@app.get("/manual-adjustment", response_class=HTMLResponse)
def manual_adjustment(request: Request, week: int = 1, day: str = "Mon", period: str | None = None):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        ensure_duty_event_rows(conn)
        assignments = {
            f"{row['week']}:{row['day']}:{row['period']}": row["staff_initials"] or ""
            for row in conn.execute("SELECT week, day, period, staff_initials FROM rota_assignments").fetchall()
        }
        staff = []
        for row in conn.execute(
            """
            SELECT initials, COALESCE(full_name, initials) AS name, classification AS role
            FROM teachers
            ORDER BY initials
            """
        ).fetchall():
            staff.append({"initials": row["initials"], "name": row["name"], "role": row["role"]})
        for row in conn.execute(
            """
            SELECT initials, full_name AS name, category AS role
            FROM additional_staff
            WHERE COALESCE(is_archived, 0) = 0
            ORDER BY category, initials
            """
        ).fetchall():
            staff.append({"initials": row["initials"], "name": row["name"], "role": row["role"]})
        duty_sections = active_duty_sections(conn)
    return templates.TemplateResponse(
        "manual_adjustment.html",
        _base_context(
            request,
            "Manual Adjustment",
            duty_sections=duty_sections,
            duty_labels=DUTY_LABELS,
            assignments=assignments,
            selected={"week": week, "day": day, "period": period},
            staff=staff,
        ),
    )


@app.post("/manual-adjustment/assign")
async def manual_adjustment_assign(
    request: Request,
    week: int = Form(...),
    day: str = Form(...),
    period: str = Form(...),
    staff_value: str = Form(...),
):
    redirect = _require_login(request)
    if redirect:
        return redirect
    initials, role = staff_value.split("|", 1)
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Manual adjustment autosave")
        cleared = assign_staff(conn, week, day, period, initials, role, enforce_rules=False, assignment_source="manual_adjustment")
    if cleared < 0:
        _flash(request, "That manual assignment would break an active rule, so the duty was left unchanged.")
        return RedirectResponse(f"/manual-adjustment?week={week}&day={day}&period={period}", status_code=303)
    note = f" Cleared {cleared} Period 4 clash(es)." if cleared else ""
    _flash(request, f"Manual adjustment saved for {initials}.{note}")
    return RedirectResponse(f"/manual-adjustment?week={week}&day={day}&period={period}", status_code=303)


@app.post("/manual-adjustment/clear")
async def manual_adjustment_clear(request: Request, week: int = Form(...), day: str = Form(...), period: str = Form(...)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Manual adjustment autosave")
        conn.execute(
            "UPDATE rota_assignments SET staff_initials = NULL, staff_type = NULL, assignment_source = NULL, last_updated = ? WHERE week = ? AND day = ? AND period = ?",
            (datetime.now().isoformat(), week, day, period),
        )
        conn.commit()
    _flash(request, "Manual assignment cleared.")
    return RedirectResponse(f"/manual-adjustment?week={week}&day={day}&period={period}", status_code=303)


@app.get("/staff-timetable", response_class=HTMLResponse)
def staff_timetable(request: Request, group: str = "Teacher", staff: str = ""):
    redirect = _require_login(request)
    if redirect:
        return redirect
    group_options = ["All", "Teacher", "HOF", "SLT", "Pastoral", "ESLT", "Chaplaincy", "Admin"]
    if group not in group_options:
        group = "Teacher"
    with _conn_context() as conn:
        people = []
        for row in conn.execute(
            """
            SELECT initials, COALESCE(full_name, initials) AS name, classification AS role, 'teacher' AS source
            FROM teachers
            ORDER BY classification, initials
            """
        ).fetchall():
            if group in {"All", row["role"]}:
                people.append(dict(row))
        for row in conn.execute(
            """
            SELECT initials, full_name AS name, category AS role, 'additional' AS source
            FROM additional_staff
            WHERE COALESCE(is_archived, 0) = 0
            ORDER BY category, initials
            """
        ).fetchall():
            if group in {"All", row["role"]}:
                people.append(dict(row))
        if not staff and people:
            staff = people[0]["initials"]
        selected = next((person for person in people if person["initials"] == staff), None)
        timetable = {}
        duty_rows = []
        teaching_rows = []
        for week in ROTA_WEEKS:
            for day in ROTA_DAYS:
                for code, _label in STAFF_TIMETABLE_COLUMNS:
                    timetable[f"{week}:{day}:{code}"] = []
        if selected:
            if selected["source"] == "teacher":
                teaching_rows = conn.execute(
                    """
                    SELECT week, day, period
                    FROM teacher_periods
                    WHERE teacher_initials = ?
                    ORDER BY week, day, period
                    """,
                    (staff,),
                ).fetchall()
                for row in teaching_rows:
                    key = f"{row['week']}:{row['day']}:{row['period']}"
                    if key in timetable:
                        timetable[key].append({"kind": "Teaching", "label": "Teaching"})
            duty_rows = conn.execute(
                """
                SELECT week, day, period, staff_type
                FROM rota_assignments
                WHERE staff_initials = ?
                ORDER BY week,
                         CASE day WHEN 'Mon' THEN 1 WHEN 'Tue' THEN 2 WHEN 'Wed' THEN 3 WHEN 'Thu' THEN 4 ELSE 5 END,
                         period
                """,
                (staff,),
            ).fetchall()
            for row in duty_rows:
                group_code = _staff_timetable_group_for_duty(row["period"])
                key = f"{row['week']}:{row['day']}:{group_code}"
                if key in timetable:
                    timetable[key].append({"kind": "Duty", "label": DUTY_LABELS.get(row["period"], row["period"])})
        summary = {"teaching": len(teaching_rows), "duties": len(duty_rows)}
    return templates.TemplateResponse(
        "staff_timetable.html",
        _base_context(
            request,
            "Staff Timetable",
            group_options=group_options,
            selected_group=group,
            people=people,
            selected=selected,
            selected_staff=staff,
            columns=STAFF_TIMETABLE_COLUMNS,
            timetable=timetable,
            summary=summary,
        ),
    )


@app.get("/proposed", response_class=HTMLResponse)
def proposed(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        active_codes = active_duty_codes(conn)
        assignments = conn.execute(
            """
            SELECT week, day, period, staff_initials, staff_type FROM rota_assignments
            ORDER BY week,
                     CASE day WHEN 'Mon' THEN 1 WHEN 'Tue' THEN 2 WHEN 'Wed' THEN 3 WHEN 'Thu' THEN 4 ELSE 5 END
            """
        ).fetchall()
        assignments = [row for row in assignments if row["period"] in active_codes]
        review = proposed_review(conn)
        blank_reasons = {f"{row['week']}:{row['day']}:{row['period']}": row["reason"] for row in review["blanks"]}
    return templates.TemplateResponse(
        "proposed.html",
        _base_context(
            request,
            "Proposed Timetable",
            assignments=assignments,
            duty_labels=DUTY_LABELS,
            duty_order=DUTY_ORDER,
            review=review,
            blank_reasons=blank_reasons,
        ),
    )


@app.get("/proposed/download")
def proposed_download(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_version_snapshot(conn, f"Proposed timetable {datetime.now().strftime('%d %b %H:%M')}", "Saved when proposed timetable was downloaded")
        data = build_master_style_workbook(conn)
    return Response(
        content=data.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Proposed_Timetable_Master_Duty_Format.xlsx"},
    )


@app.get("/versions", response_class=HTMLResponse)
def versions(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        rows = conn.execute(
            """
            SELECT id, name, reason, created_at
            FROM versions
            ORDER BY id DESC
            LIMIT 80
            """
        ).fetchall()
    return templates.TemplateResponse("versions.html", _base_context(request, "Versions", rows=rows))


@app.post("/versions/create")
async def versions_create(request: Request, name: str = Form("Manual save"), reason: str = Form("Manual version save")):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_version_snapshot(conn, name, reason.strip() or "Manual version save")
    _flash(request, "Version saved.")
    return RedirectResponse("/versions", status_code=303)


@app.post("/versions/restore")
async def versions_restore(request: Request, version_id: int = Form(...)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        restored_name = restore_version_snapshot(conn, version_id)
    _flash(request, f"Restored version: {restored_name}. Additional staff added later were kept.")
    return RedirectResponse("/versions", status_code=303)


@app.post("/versions/autosave")
async def versions_autosave(request: Request):
    if not _is_logged_in(request):
        return Response(status_code=204)
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Page leave autosave")
    return Response(status_code=204)


@app.get("/problems", response_class=HTMLResponse)
def problems(request: Request):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        rows = conn.execute("SELECT * FROM problem_log ORDER BY timestamp DESC").fetchall()
    return templates.TemplateResponse("problems.html", _base_context(request, "Narrative & Problems", rows=rows))


@app.post("/problems/add")
async def problems_add(request: Request, issue_type: str = Form(...), description: str = Form(...)):
    redirect = _require_login(request)
    if redirect:
        return redirect
    with _conn_context() as conn:
        create_throttled_autosave(conn, "Narrative edit autosave")
        conn.execute(
            "INSERT INTO problem_log(issue_type, description) VALUES (?, ?)",
            (issue_type, description.strip()),
        )
        conn.commit()
    return RedirectResponse("/problems", status_code=303)
