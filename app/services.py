from __future__ import annotations

import re
import json
import sqlite3
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .constants import (
    DUTY_LABELS,
    DUTY_ORDER,
    DUTY_SECTIONS,
    EXPORTS_DIR,
    MASTER_TEMPLATE_PATH,
    ROTA_DAYS,
    ROTA_WEEKS,
)


SNAPSHOT_TABLES = [
    "timetable_meta",
    "teachers",
    "teacher_periods",
    "staff_names",
    "classifications",
    "additional_staff",
    "rota_assignments",
    "rules",
    "custom_rules",
    "app_settings",
    "staff_exclusions",
    "problem_log",
]

FULL_TIME_TEACHING_LOAD = 65.0
DAILY_TEACHING_LOAD = 6.5


def parse_timetable(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb.active
    school_name = ws.cell(row=3, column=2).value or "Unknown School"

    week1_label = None
    week2_label = None
    for row in range(1, 50):
        for col in range(1, 50):
            val = ws.cell(row=row, column=col).value
            if val and "Week 1" in str(val):
                week1_label = str(val)
            if val and "Week 2" in str(val):
                week2_label = str(val)

    day_map = {
        "Monday": "Mon",
        "Tuesday": "Tue",
        "Wednesday": "Wed",
        "Thursday": "Thu",
        "Friday": "Fri",
    }
    teaching_periods = {"Tutor", "1", "2", "3", "4", "5", "6"}
    all_period_headers = []
    current_day = None
    current_week = 1

    for col in range(1, ws.max_column + 1):
        day_val = ws.cell(row=13, column=col).value
        period_val = ws.cell(row=14, column=col).value
        if day_val in day_map:
            if day_val == "Monday" and current_day == "Fri":
                current_week = 2
            current_day = day_map[day_val]
        if current_day and period_val:
            all_period_headers.append(
                {"week": current_week, "day": current_day, "period": str(period_val), "col": col}
            )

    period_columns = []
    for idx, header in enumerate(all_period_headers):
        if header["period"] in teaching_periods:
            next_col = all_period_headers[idx + 1]["col"] if idx + 1 < len(all_period_headers) else ws.max_column + 1
            period_columns.append({**header, "end_col": next_col - 1})

    teacher_period_rows = []
    teacher_subject_counts: dict[str, dict[str, int]] = {}
    real_teacher_day_presence: dict[str, set[tuple[int, str]]] = {}
    seen_period_rows = set()
    for period_info in period_columns:
        for row in range(15, ws.max_row + 1):
            for col in range(period_info["col"], period_info["end_col"] + 1):
                val = ws.cell(row=row, column=col).value
                if val and isinstance(val, str):
                    initials = val.strip()
                    if re.match(r"^[A-Z]{3}$", initials):
                        subject = infer_subject_for_cell(ws, row, col)
                        if subject:
                            subject_counts = teacher_subject_counts.setdefault(initials, {})
                            subject_counts[subject] = subject_counts.get(subject, 0) + 1
                        key = (initials, period_info["week"], period_info["day"], period_info["period"])
                        if key not in seen_period_rows:
                            seen_period_rows.add(key)
                            real_teacher_day_presence.setdefault(initials, set()).add((period_info["week"], period_info["day"]))
                            teacher_period_rows.append(
                                {
                                    "teacher_initials": initials,
                                    "week": period_info["week"],
                                    "day": period_info["day"],
                                    "period": period_info["period"],
                                    "source_row": row,
                                    "source_col": col,
                                }
                            )

    teachers_seen = {row["teacher_initials"] for row in teacher_period_rows}
    occupied_slots = {
        (row["teacher_initials"], row["week"], row["day"], row["period"])
        for row in teacher_period_rows
    }
    for initials in sorted(teachers_seen):
        for week in [1, 2]:
            if (week, "Mon") in real_teacher_day_presence.get(initials, set()) and (initials, week, "Mon", "6") not in occupied_slots:
                teacher_period_rows.append(
                    {
                        "teacher_initials": initials,
                        "week": week,
                        "day": "Mon",
                        "period": "6",
                        "source_row": -1,
                        "source_col": -1,
                    }
                )

    teacher_counts: dict[str, float] = {}
    for row in teacher_period_rows:
        weight = 0.5 if row["period"] == "Tutor" else 1
        teacher_counts[row["teacher_initials"]] = teacher_counts.get(row["teacher_initials"], 0) + weight

    day_order = [(1, "Mon"), (1, "Tue"), (1, "Wed"), (1, "Thu"), (1, "Fri"),
                 (2, "Mon"), (2, "Tue"), (2, "Wed"), (2, "Thu"), (2, "Fri")]
    teachers_data = []
    for initials in sorted(teacher_counts):
        total = teacher_counts[initials]
        days_in_school = "".join(
            "1" if day_key in real_teacher_day_presence.get(initials, set()) else "0"
            for day_key in day_order
        )
        days_out = days_in_school.count("0")
        max_load = FULL_TIME_TEACHING_LOAD - (DAILY_TEACHING_LOAD * days_out)
        subject_counts = teacher_subject_counts.get(initials, {})
        detected_subject = max(subject_counts, key=subject_counts.get) if subject_counts else ""
        teachers_data.append(
            {
                "initials": initials,
                "full_name": f"Teacher {initials}",
                "is_teaching": 1 if total > 0 else 0,
                "lessons_week1": 0,
                "lessons_week2": total,
                "total_lessons": total,
                "non_contact": max(0, max_load - total),
                "protected_periods": 6,
                "classification": "Teacher",
                "is_part_time": 1 if "0" in days_in_school else 0,
                "days_in_school": days_in_school,
                "subject": detected_subject,
                "max_lunch_duties": None,
                "exclude_from_algorithm": 0,
            }
        )

    return {
        "school_name": school_name,
        "week1_label": week1_label or "Week 1",
        "week2_label": week2_label or "Week 2 (inferred)",
        "teachers": teachers_data,
        "teacher_periods": teacher_period_rows,
    }


def infer_subject_for_cell(ws, row: int, col: int) -> str:
    ignored = {"", "AM", "PM", "BR", "BR1", "Tutor", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"}
    for look_col in range(max(1, col - 5), col):
        value = ws.cell(row=row, column=look_col).value
        text = str(value or "").strip()
        if 2 <= len(text) <= 40 and text not in ignored and not re.match(r"^[A-Z]{3}$", text):
            return text
    for look_row in range(max(1, row - 8), row):
        value = ws.cell(row=look_row, column=1).value or ws.cell(row=look_row, column=2).value
        text = str(value or "").strip()
        if 2 <= len(text) <= 40 and text not in ignored and not re.match(r"^[A-Z]{3}$", text):
            return text
    return ""


def table_rows_as_dicts(conn: sqlite3.Connection, table_name: str) -> list[dict]:
    return [dict(row) for row in conn.execute(f"SELECT * FROM {table_name}").fetchall()]


def create_version_snapshot(conn: sqlite3.Connection, name: str, reason: str = "Manual snapshot") -> int:
    snapshot = {table: table_rows_as_dicts(conn, table) for table in SNAPSHOT_TABLES}
    cursor = conn.execute(
        """
        INSERT INTO versions(name, reason, snapshot_json, created_at)
        VALUES (?, ?, ?, ?)
        """,
        (name.strip() or "DutyMaster snapshot", reason, json.dumps(snapshot), datetime.now().isoformat()),
    )
    conn.commit()
    return int(cursor.lastrowid)


def create_throttled_autosave(conn: sqlite3.Connection, reason: str = "Autosave") -> int | None:
    latest = conn.execute(
        """
        SELECT created_at FROM versions
        WHERE reason = ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (reason,),
    ).fetchone()
    if latest:
        try:
            previous = datetime.fromisoformat(latest["created_at"])
            if (datetime.now() - previous).total_seconds() < 120:
                return None
        except ValueError:
            pass
    return create_version_snapshot(conn, f"Autosave {datetime.now().strftime('%d %b %H:%M')}", reason)


def _insert_rows(conn: sqlite3.Connection, table_name: str, rows: list[dict]) -> None:
    if not rows:
        return
    columns = [row["name"] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()]
    for row in rows:
        usable = {key: row[key] for key in columns if key in row}
        if not usable:
            continue
        names = list(usable)
        placeholders = ", ".join("?" for _ in names)
        conn.execute(
            f"INSERT OR REPLACE INTO {table_name} ({', '.join(names)}) VALUES ({placeholders})",
            [usable[name] for name in names],
        )


def _restore_additional_staff(conn: sqlite3.Connection, rows: list[dict]) -> None:
    columns = [row["name"] for row in conn.execute("PRAGMA table_info(additional_staff)").fetchall()]
    for row in rows:
        initials = row.get("initials")
        if not initials:
            continue
        existing = conn.execute("SELECT id FROM additional_staff WHERE initials = ?", (initials,)).fetchone()
        usable = {key: row[key] for key in columns if key in row and key != "id"}
        if existing:
            assignments = ", ".join(f"{key} = ?" for key in usable)
            conn.execute(
                f"UPDATE additional_staff SET {assignments} WHERE initials = ?",
                [*usable.values(), initials],
            )
        else:
            names = list(usable)
            placeholders = ", ".join("?" for _ in names)
            conn.execute(
                f"INSERT INTO additional_staff ({', '.join(names)}) VALUES ({placeholders})",
                [usable[name] for name in names],
            )


def restore_version_snapshot(conn: sqlite3.Connection, version_id: int) -> str:
    version = conn.execute("SELECT name, snapshot_json FROM versions WHERE id = ?", (version_id,)).fetchone()
    if not version:
        raise ValueError("Version not found")
    create_version_snapshot(conn, f"Before restore {datetime.now().strftime('%d %b %H:%M')}", "Automatic safety copy before restore")
    snapshot = json.loads(version["snapshot_json"])
    for table in SNAPSHOT_TABLES:
        if table == "additional_staff":
            continue
        conn.execute(f"DELETE FROM {table}")
        _insert_rows(conn, table, snapshot.get(table, []))
    _restore_additional_staff(conn, snapshot.get("additional_staff", []))
    from .database import seed_defaults
    seed_defaults(conn)
    ensure_duty_event_rows(conn)
    conn.commit()
    return version["name"]


def reset_upload_data(conn: sqlite3.Connection) -> None:
    for table in ["timetable_meta", "teachers", "teacher_periods", "rota_assignments"]:
        conn.execute(f"DELETE FROM {table}")
    conn.commit()


def save_parsed_timetable(conn: sqlite3.Connection, parsed: dict) -> None:
    reset_upload_data(conn)
    conn.execute(
        "INSERT INTO timetable_meta (school_name, week1_label, week2_label, uploaded_at) VALUES (?, ?, ?, ?)",
        (parsed["school_name"], parsed["week1_label"], parsed["week2_label"], datetime.now().isoformat()),
    )
    for t in parsed["teachers"]:
        conn.execute(
            """
            INSERT INTO teachers (
                initials, full_name, is_teaching, lessons_week1, lessons_week2, total_lessons,
                non_contact, protected_periods, classification, is_part_time, days_in_school, subject,
                max_lunch_duties, exclude_from_algorithm, last_updated
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                t["initials"], t["full_name"], t["is_teaching"], t["lessons_week1"], t["lessons_week2"],
                t["total_lessons"], t["non_contact"], t["protected_periods"], t["classification"],
                t["is_part_time"], t["days_in_school"], t.get("subject", ""), t.get("max_lunch_duties"),
                t.get("exclude_from_algorithm", 0), datetime.now().isoformat(),
            ),
        )
    for row in parsed["teacher_periods"]:
        conn.execute(
            """
            INSERT OR IGNORE INTO teacher_periods
            (teacher_initials, week, day, period, source_row, source_col)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (row["teacher_initials"], row["week"], row["day"], row["period"], row["source_row"], row["source_col"]),
        )
    ensure_duty_event_rows(conn)
    conn.commit()


def ensure_duty_event_rows(conn: sqlite3.Connection) -> None:
    for week in ROTA_WEEKS:
        for day in ROTA_DAYS:
            for code in DUTY_LABELS:
                conn.execute(
                    "INSERT OR IGNORE INTO rota_assignments (week, day, period) VALUES (?, ?, ?)",
                    (week, day, code),
                )
    conn.commit()


def clear_inactive_teacher_break_slots(conn: sqlite3.Connection) -> int:
    active_slots = teacher_break_rota_slots(conn)
    cursor = conn.execute(
        """
        UPDATE rota_assignments
        SET staff_initials = NULL, staff_type = NULL, assignment_source = NULL, last_updated = ?
        WHERE period LIKE 'Teacher_Break_Rota_%'
          AND CAST(SUBSTR(period, 20) AS INTEGER) > ?
          AND staff_initials IS NOT NULL
        """,
        (datetime.now().isoformat(), active_slots),
    )
    conn.commit()
    return cursor.rowcount


def event_to_timetable_period(code: str) -> str | None:
    if code.startswith("Teacher_Break_Rota_"):
        return None
    if code.startswith("Tutor_"):
        return "Tutor"
    for number in ["1", "2", "3", "5", "6"]:
        if code.startswith(f"P{number}_"):
            return number
    if code.startswith("P4"):
        return "4"
    return None


def duty_time_group(code: str) -> str:
    if code == "Gate":
        return "Gate"
    for prefix, group in [
        ("Tutor_", "Tutor"), ("P1_", "P1"), ("P2_", "P2"), ("Break_", "Break"),
        ("Teacher_Break_Rota_", "TeacherBreak"),
        ("P3_", "P3"), ("P4A_", "P4A"), ("P4B_", "P4B"), ("P4C_", "P4C"),
        ("P4_", "P4_Lunch"), ("P5_", "P5"), ("P6_", "P6"), ("P7_", "P7"),
    ]:
        if code.startswith(prefix):
            return group
    return code


def event_to_availability_period(code: str) -> str | None:
    group = duty_time_group(code)
    mapping = {
        "Tutor": "Tutor",
        "P1": "1",
        "P2": "2",
        "Break": "Break",
        "TeacherBreak": "TeacherBreak",
        "P3": "3",
        "P4_Lunch": "4",
        "P4A": "4",
        "P4B": "4",
        "P4C": "4",
        "P5": "5",
        "P6": "6",
        "P7": "7",
    }
    return mapping.get(group)


def duty_family(code: str) -> str:
    family_markers = [
        ("Pastoral_Support", "Pastoral Support"),
        ("Room_90", "Room 90"),
        ("Isolation", "Isolation"),
        ("First_Duty", "First Duty"),
        ("Late_Detention", "Late Detention"),
        ("Duty_Lead", "Duty Lead"),
        ("Rest_Break", "Rest Break"),
        ("AOW", "Act of Worship"),
    ]
    for marker, family in family_markers:
        if marker in code:
            return family
    if code.startswith("P4_Lunch_"):
        return "Lunch Duty"
    if code.startswith("P7_Detention"):
        return "Detention Duty"
    return code


def previous_duty_time_groups(group: str) -> list[str]:
    sequence = ["Tutor", "P1", "P2", "Break", "TeacherBreak", "P3", "P4A", "P4B", "P4C", "P5", "P6", "P7"]
    if group == "P4_Lunch":
        return ["P3"]
    if group not in sequence:
        return []
    index = sequence.index(group)
    if index == 0:
        return []
    previous = sequence[index - 1]
    if previous == "TeacherBreak":
        return ["TeacherBreak", "Break"]
    return [previous]


def pastoral_repeats_same_duty_previous_period(conn: sqlite3.Connection, initials: str, week: int, day: str, code: str) -> bool:
    current_family = duty_family(code)
    previous_groups = set(previous_duty_time_groups(duty_time_group(code)))
    if not previous_groups:
        return False
    rows = conn.execute(
        """
        SELECT period
        FROM rota_assignments
        WHERE staff_initials = ? AND week = ? AND day = ?
        """,
        (initials, week, day),
    ).fetchall()
    return any(
        duty_time_group(row["period"]) in previous_groups
        and duty_family(row["period"]) == current_family
        for row in rows
    )


def groups_conflict(existing_group: str, requested_group: str) -> bool:
    if existing_group == requested_group:
        return True
    p4_phases = {"P4A", "P4B", "P4C"}
    return (
        existing_group == "P4_Lunch" and requested_group in p4_phases
    ) or (
        requested_group == "P4_Lunch" and existing_group in p4_phases
    )


def same_time_assignment_exists(conn: sqlite3.Connection, initials: str, week: int, day: str, code: str) -> bool:
    current_group = duty_time_group(code)
    rows = conn.execute(
        """
        SELECT period FROM rota_assignments
        WHERE staff_initials = ? AND week = ? AND day = ?
        """,
        (initials, week, day),
    ).fetchall()
    return any(groups_conflict(duty_time_group(row["period"]), current_group) for row in rows)


def teacher_available(conn: sqlite3.Connection, initials: str, week: int, day: str, code: str) -> bool:
    row = conn.execute(
        "SELECT days_in_school, COALESCE(exclude_from_algorithm, 0) AS exclude_from_algorithm FROM teachers WHERE initials = ?",
        (initials,),
    ).fetchone()
    if not row:
        return False
    if row["exclude_from_algorithm"]:
        return False
    days = (row["days_in_school"] or "1111111111").ljust(10, "1")[:10]
    idx = (week - 1) * 5 + ROTA_DAYS.index(day)
    if days[idx] != "1":
        return False
    teaching_period = event_to_timetable_period(code)
    if teaching_period:
        busy = conn.execute(
            """
            SELECT 1 FROM teacher_periods
            WHERE teacher_initials = ? AND week = ? AND day = ? AND period = ?
            LIMIT 1
            """,
            (initials, week, day, teaching_period),
        ).fetchone()
        if busy:
            return False
    return not same_time_assignment_exists(conn, initials, week, day, code)


def additional_available(conn: sqlite3.Connection, initials: str, week: int, day: str, code: str) -> bool:
    row = conn.execute(
        """
        SELECT status, COALESCE(days_in_school, '1111111111') AS days_in_school,
               COALESCE(availability, '[]') AS availability
        FROM additional_staff
        WHERE initials = ? AND COALESCE(is_archived, 0) = 0
        """,
        (initials,),
    ).fetchone()
    if not row or (row["status"] or "Active") != "Active":
        return False
    days = (row["days_in_school"] or "1111111111").ljust(10, "1")[:10]
    idx = (week - 1) * 5 + ROTA_DAYS.index(day)
    if days[idx] != "1":
        return False
    try:
        periods = set(json.loads(row["availability"] or "[]"))
    except (TypeError, json.JSONDecodeError):
        periods = set()
    requested_period = event_to_availability_period(code)
    if requested_period and requested_period not in periods:
        return False
    return not same_time_assignment_exists(conn, initials, week, day, code)


def role_priority(role: str) -> int:
    return {"Pastoral": 0, "SLT": 1, "HOF": 2, "Teacher": 3, "ESLT": 4, "Chaplaincy": 5, "Admin": 6}.get(role, 9)


def duty_is_heavy(code: str) -> bool:
    return "Isolation" in code or "Lunch" in code or "Detention" in code


def duty_is_lunch(code: str) -> bool:
    return code.startswith("P4_Lunch_")


def p7_mode(conn: sqlite3.Connection) -> str:
    mode = get_setting(conn, "p7_mode", get_setting(conn, "p7_detention_mode", "ignore")).strip().lower()
    return mode if mode in {"ignore", "slt", "pastoral"} else "ignore"


def duty_is_optional(code: str) -> bool:
    return "Room_90" in code


def duty_is_optional_for_conn(conn: sqlite3.Connection, code: str) -> bool:
    return duty_is_optional(code) or (code.startswith("P7_") and p7_mode(conn) == "ignore")


def duty_is_manual_fill_only_for_conn(conn: sqlite3.Connection, code: str) -> bool:
    return "Room_90" in code and rule_active(conn, "Room 90 Manual Fill Only")


def get_setting(conn: sqlite3.Connection, key: str, default: str) -> str:
    row = conn.execute("SELECT value FROM app_settings WHERE key = ?", (key,)).fetchone()
    return str(row["value"]) if row else default


def max_duties_per_week(conn: sqlite3.Connection) -> int:
    try:
        return int(float(get_setting(conn, "max_duties_per_week", "4")))
    except ValueError:
        return 4


def max_duties_per_day(conn: sqlite3.Connection) -> int:
    try:
        return int(float(get_setting(conn, "max_duties_per_day", "2")))
    except ValueError:
        return 2


def teacher_break_rota_slots(conn: sqlite3.Connection) -> int:
    try:
        return max(0, min(10, int(float(get_setting(conn, "teacher_break_rota_slots", "6")))))
    except ValueError:
        return 6


def rule_active(conn: sqlite3.Connection, name: str, default: bool = True) -> bool:
    row = conn.execute("SELECT active FROM rules WHERE name = ?", (name,)).fetchone()
    if not row:
        return default
    return bool(row["active"])


def active_duty_sections(conn: sqlite3.Connection) -> list[tuple[str, list[tuple[str, str]]]]:
    active_slots = teacher_break_rota_slots(conn)
    sections = []
    for section, events in DUTY_SECTIONS:
        if section == "Teaching Staff Break Rota":
            events = [(code, label) for code, label in events if int(code.rsplit("_", 1)[1]) <= active_slots]
        if section == "Period 7" and p7_mode(conn) == "ignore":
            continue
        sections.append((section, events))
    return sections


def active_duty_codes(conn: sqlite3.Connection) -> set[str]:
    return {code for _, events in active_duty_sections(conn) for code, _ in events}


def duty_scope_matches(code: str, scope: str) -> bool:
    scope = scope or "Any"
    if scope == "Any":
        return True
    if scope == code:
        return True
    scope_prefixes = {
        "Gate Duty": ["Gate"],
        "Tutor Time": ["Tutor_"],
        "Period 1": ["P1_"],
        "Period 2": ["P2_"],
        "Break": ["Break_"],
        "Teacher Break Rota": ["Teacher_Break_Rota_"],
        "Teaching Staff Break Rota": ["Teacher_Break_Rota_"],
        "Period 3": ["P3_"],
        "Period 4 Lunch": ["P4_Lunch_"],
        "Period 4A": ["P4A_"],
        "Period 4B": ["P4B_"],
        "Period 4C": ["P4C_"],
        "Period 5": ["P5_"],
        "Period 6": ["P6_"],
        "Period 7": ["P7_"],
        "Isolation Duties": ["Tutor_Isolation", "P1_Isolation", "P2_Isolation", "Break_Isolation", "P3_Isolation", "P4A_Isolation", "P4B_Isolation", "P4C_Isolation", "P5_Isolation", "P6_Isolation"],
        "Lunch and Detention": ["P4_Lunch_", "Break_Late_Detention", "P7_Detention"],
        "Heavy Duties": ["Isolation", "Lunch", "Detention"],
    }
    prefixes = scope_prefixes.get(scope, [])
    return any(code.startswith(prefix) or prefix in code for prefix in prefixes)


def staff_scope_matches(initials: str, role: str, scope: str) -> bool:
    scope = scope or "Any"
    if scope == "Any":
        return True
    if scope.startswith("Staff:"):
        return initials == scope.split(":", 1)[1]
    return role == scope


def staff_week_duty_count(conn: sqlite3.Connection, initials: str, week: int) -> int:
    row = conn.execute(
        "SELECT COUNT(*) AS count FROM rota_assignments WHERE staff_initials = ? AND week = ?",
        (initials, week),
    ).fetchone()
    return int(row["count"] or 0)


def staff_day_duty_count(conn: sqlite3.Connection, initials: str, week: int, day: str) -> int:
    row = conn.execute(
        "SELECT COUNT(*) AS count FROM rota_assignments WHERE staff_initials = ? AND week = ? AND day = ?",
        (initials, week, day),
    ).fetchone()
    return int(row["count"] or 0)


def staff_total_duty_count(conn: sqlite3.Connection, initials: str) -> int:
    row = conn.execute(
        "SELECT COUNT(*) AS count FROM rota_assignments WHERE staff_initials = ?",
        (initials,),
    ).fetchone()
    return int(row["count"] or 0)


def staff_lunch_duty_count(conn: sqlite3.Connection, initials: str) -> int:
    row = conn.execute(
        "SELECT COUNT(*) AS count FROM rota_assignments WHERE staff_initials = ? AND period LIKE 'P4_Lunch_%'",
        (initials,),
    ).fetchone()
    return int(row["count"] or 0)


def teacher_free_periods_remaining(conn: sqlite3.Connection, initials: str) -> float:
    row = conn.execute(
        "SELECT total_lessons, protected_periods, days_in_school FROM teachers WHERE initials = ?",
        (initials,),
    ).fetchone()
    if not row:
        return 0
    days = (row["days_in_school"] or "1111111111").ljust(10, "1")[:10]
    max_load = FULL_TIME_TEACHING_LOAD - (DAILY_TEACHING_LOAD * days.count("0"))
    return max_load - float(row["total_lessons"] or 0) - int(row["protected_periods"] or 0) - staff_total_duty_count(conn, initials)


def teacher_free_periods_for_day(conn: sqlite3.Connection, initials: str, week: int, day: str) -> float:
    teaching_rows = conn.execute(
        """
        SELECT period FROM teacher_periods
        WHERE teacher_initials = ? AND week = ? AND day = ?
        """,
        (initials, week, day),
    ).fetchall()
    teaching = sum(0.5 if row["period"] == "Tutor" else 1 for row in teaching_rows)
    return DAILY_TEACHING_LOAD - teaching - staff_day_duty_count(conn, initials, week, day)


def teacher_lunch_limit_reached(conn: sqlite3.Connection, initials: str) -> bool:
    row = conn.execute("SELECT max_lunch_duties FROM teachers WHERE initials = ?", (initials,)).fetchone()
    if not row or row["max_lunch_duties"] is None:
        return False
    return staff_lunch_duty_count(conn, initials) >= int(row["max_lunch_duties"])


def additional_max_reached(conn: sqlite3.Connection, initials: str) -> bool:
    row = conn.execute("SELECT max_duties FROM additional_staff WHERE initials = ?", (initials,)).fetchone()
    if not row or row["max_duties"] is None:
        return False
    return staff_total_duty_count(conn, initials) >= int(row["max_duties"])


def teacher_break_week_count(conn: sqlite3.Connection, initials: str, week: int) -> int:
    row = conn.execute(
        """
        SELECT COUNT(*) AS count
        FROM rota_assignments
        WHERE staff_initials = ? AND week = ? AND period LIKE 'Teacher_Break_Rota_%'
        """,
        (initials, week),
    ).fetchone()
    return int(row["count"] or 0)


def assigned_teacher_break_subjects(conn: sqlite3.Connection, week: int, day: str) -> set[str]:
    rows = conn.execute(
        """
        SELECT COALESCE(t.subject, '') AS subject
        FROM rota_assignments r
        JOIN teachers t ON t.initials = r.staff_initials
        WHERE r.week = ? AND r.day = ? AND r.period LIKE 'Teacher_Break_Rota_%'
          AND COALESCE(t.subject, '') != ''
        """,
        (week, day),
    ).fetchall()
    return {row["subject"] for row in rows}


def previous_non_free_streak(conn: sqlite3.Connection, initials: str, week: int, day: str, code: str) -> int:
    ordered_slots = []
    for current_week in ROTA_WEEKS:
        for current_day in ROTA_DAYS:
            for duty_code in DUTY_ORDER:
                ordered_slots.append((current_week, current_day, duty_code))
    current = (week, day, code)
    if current not in ordered_slots:
        return 0
    index = ordered_slots.index(current)
    streak = 0
    for prev_week, prev_day, prev_code in reversed(ordered_slots[:index]):
        timetable_period = event_to_timetable_period(prev_code)
        taught = False
        if timetable_period:
            taught = bool(
                conn.execute(
                    """
                    SELECT 1 FROM teacher_periods
                    WHERE teacher_initials = ? AND week = ? AND day = ? AND period = ?
                    LIMIT 1
                    """,
                    (initials, prev_week, prev_day, timetable_period),
                ).fetchone()
            )
        duty = bool(
            conn.execute(
                """
                SELECT 1 FROM rota_assignments
                WHERE staff_initials = ? AND week = ? AND day = ? AND period = ?
                LIMIT 1
                """,
                (initials, prev_week, prev_day, prev_code),
            ).fetchone()
        )
        if taught or duty:
            streak += 1
            continue
        break
    return streak


def teacher_assignment_score(conn: sqlite3.Connection, initials: str, week: int, day: str, code: str) -> tuple[float, float]:
    remaining = teacher_free_periods_remaining(conn, initials)
    score = 100 + (10 * remaining)
    if teacher_free_periods_for_day(conn, initials, week, day) <= 1:
        score -= 1
    tie_break = score - previous_non_free_streak(conn, initials, week, day, code)
    return score, tie_break


def staff_excluded_on_day(conn: sqlite3.Connection, initials: str, week: int, day: str) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT reason
        FROM staff_exclusions
        WHERE staff_initials = ? AND week = ? AND day = ? AND active = 1
        ORDER BY id DESC
        LIMIT 1
        """,
        (initials, week, day),
    ).fetchone()


def has_heavy_duty_same_day(conn: sqlite3.Connection, initials: str, week: int, day: str) -> bool:
    rows = conn.execute(
        "SELECT period FROM rota_assignments WHERE staff_initials = ? AND week = ? AND day = ?",
        (initials, week, day),
    ).fetchall()
    return any(duty_is_heavy(row["period"]) for row in rows)


def teacher_available_periods(conn: sqlite3.Connection, initials: str) -> float:
    row = conn.execute(
        """
        SELECT total_lessons, protected_periods, days_in_school
        FROM teachers
        WHERE initials = ?
        """,
        (initials,),
    ).fetchone()
    if not row:
        return 999
    days = (row["days_in_school"] or "1111111111").ljust(10, "1")[:10]
    max_load = FULL_TIME_TEACHING_LOAD - (DAILY_TEACHING_LOAD * days.count("0"))
    current_duties = conn.execute(
        "SELECT COUNT(*) AS count FROM rota_assignments WHERE staff_initials = ?",
        (initials,),
    ).fetchone()["count"]
    return max_load - float(row["total_lessons"] or 0) - int(row["protected_periods"] or 0) - int(current_duties or 0)


def custom_rules_allow(conn: sqlite3.Connection, initials: str, role: str, week: int, day: str, code: str) -> bool:
    rules = conn.execute(
        """
        SELECT duty_scope, staff_scope, condition_type, condition_value, priority
        FROM custom_rules
        WHERE active = 1 AND COALESCE(is_archived, 0) = 0
        """
    ).fetchall()
    for rule in rules:
        if (rule["priority"] or "Hard") != "Hard":
            continue
        if not duty_scope_matches(code, rule["duty_scope"]):
            continue
        if not staff_scope_matches(initials, role, rule["staff_scope"]):
            continue
        condition = rule["condition_type"]
        value = (rule["condition_value"] or "").strip()
        if condition == "exclude":
            return False
        if condition == "min_available_periods":
            try:
                minimum = float(value)
            except ValueError:
                minimum = 0
            if teacher_available_periods(conn, initials) < minimum:
                return False
        elif condition == "max_duties_per_week":
            try:
                maximum = int(float(value))
            except ValueError:
                maximum = 0
            if staff_week_duty_count(conn, initials, week) >= maximum:
                return False
        elif condition == "no_heavy_same_day" and duty_is_heavy(code):
            if has_heavy_duty_same_day(conn, initials, week, day):
                return False
    return True


def role_priority_for_duty(code: str) -> list[str]:
    if code.startswith("Teacher_Break_Rota_"):
        return ["Teacher", "HOF", "SLT"]
    if code == "Gate":
        return ["SLT"]
    if code in {"Tutor_1st_Duty", "Tutor_AOW", "P1_Isolation"}:
        return ["SLT"]
    if code in {"P4A_First_Duty", "P4B_First_Duty", "P4C_First_Duty"}:
        return ["SLT"]
    if code in {
        "P4A_Pastoral_Support", "P4A_Isolation", "P4A_Rest_Break",
        "P4B_Pastoral_Support", "P4B_Isolation", "P4B_Rest_Break",
        "P4C_Pastoral_Support", "P4C_Isolation", "P4C_Rest_Break",
    }:
        return ["Pastoral"]
    if code == "Break_Duty_Lead":
        return ["SLT", "Pastoral"]
    if code == "Break_Late_Detention":
        return ["Pastoral"]
    if "Pastoral_Support" in code or "Room_90" in code:
        return ["Pastoral", "SLT"]
    if "Isolation" in code:
        return ["Pastoral", "SLT"]
    if "First_Duty" in code:
        return ["Pastoral", "SLT"]
    if code == "P4_Lunch_1":
        return ["Pastoral"]
    if code.startswith("P4_Lunch_"):
        return ["Teacher", "HOF", "ESLT", "Chaplaincy", "Admin", "SLT"]
    if code.startswith("P7_Detention"):
        return ["Pastoral", "SLT"]
    return ["Pastoral", "SLT", "Teacher", "HOF", "ESLT", "Chaplaincy", "Admin"]


def role_priority_for_duty_with_settings(conn: sqlite3.Connection, code: str) -> list[str]:
    if code.startswith("P7_"):
        mode = p7_mode(conn)
        if mode == "ignore":
            return []
        if mode == "slt":
            return ["SLT"]
        if mode == "pastoral":
            return ["Pastoral"]
    return role_priority_for_duty(code)


def slt_balance_sort_key(conn: sqlite3.Connection, initials: str, week: int, day: str, code: str) -> tuple:
    rows = conn.execute("SELECT period FROM rota_assignments WHERE staff_initials = ?", (initials,)).fetchall()
    heavy_count = sum(1 for row in rows if duty_is_heavy(row["period"]))
    return (
        staff_total_duty_count(conn, initials),
        heavy_count,
        previous_non_free_streak(conn, initials, week, day, code),
        initials,
    )


def duty_uses_pastoral_distribution(code: str) -> bool:
    return duty_family(code) in {"Pastoral Support", "Isolation", "Late Detention", "Rest Break"}


def pastoral_distribution_sort_key(conn: sqlite3.Connection, initials: str, week: int, day: str, code: str) -> tuple:
    rows = conn.execute(
        "SELECT period FROM rota_assignments WHERE staff_initials = ?",
        (initials,),
    ).fetchall()
    family = duty_family(code)
    pastoral_families = {"Pastoral Support", "Isolation", "Late Detention", "Rest Break"}
    same_family_count = sum(1 for row in rows if duty_family(row["period"]) == family)
    pastoral_duty_count = sum(1 for row in rows if duty_family(row["period"]) in pastoral_families)
    return (
        same_family_count,
        pastoral_duty_count,
        staff_week_duty_count(conn, initials, week),
        staff_day_duty_count(conn, initials, week, day),
        initials,
    )


def strict_assignment_allowed(
    conn: sqlite3.Connection,
    initials: str,
    role: str,
    week: int,
    day: str,
    code: str,
    source: str | None = None,
) -> bool:
    if staff_excluded_on_day(conn, initials, week, day):
        return False
    allowed_roles = role_priority_for_duty_with_settings(conn, code)
    if role not in allowed_roles:
        return False
    if role == "Pastoral" and rule_active(conn, "No Consecutive Same Pastoral Duty"):
        if pastoral_repeats_same_duty_previous_period(conn, initials, week, day, code):
            return False
    if source is None:
        source = "teacher" if conn.execute("SELECT 1 FROM teachers WHERE initials = ?", (initials,)).fetchone() else "additional"
    if source == "teacher":
        if not teacher_available(conn, initials, week, day, code):
            return False
        if duty_is_lunch(code) and teacher_lunch_limit_reached(conn, initials):
            return False
    else:
        if not additional_available(conn, initials, week, day, code):
            return False
        if role in {"ESLT", "Chaplaincy", "Admin"} and not duty_is_lunch(code):
            return False
        if additional_max_reached(conn, initials):
            return False
    if source == "teacher":
        if staff_week_duty_count(conn, initials, week) >= max_duties_per_week(conn):
            return False
        if staff_day_duty_count(conn, initials, week, day) >= max_duties_per_day(conn):
            return False
    if code.startswith("Teacher_Break_Rota_"):
        if source != "teacher":
            return False
        if teacher_break_week_count(conn, initials, week) >= 1:
            return False
    return custom_rules_allow(conn, initials, role, week, day, code)


def candidate_rejection_reason(
    conn: sqlite3.Connection,
    initials: str,
    role: str,
    week: int,
    day: str,
    code: str,
    source: str | None = None,
) -> str | None:
    exclusion = staff_excluded_on_day(conn, initials, week, day)
    if exclusion:
        return f"excluded that day: {exclusion['reason'] or 'one-off exclusion'}"
    allowed_roles = role_priority_for_duty_with_settings(conn, code)
    if role not in allowed_roles:
        return f"role {role} is not allowed for this duty"
    if role == "Pastoral" and rule_active(conn, "No Consecutive Same Pastoral Duty"):
        if pastoral_repeats_same_duty_previous_period(conn, initials, week, day, code):
            return "pastoral staff should not repeat the same duty in consecutive periods"
    if source is None:
        source = "teacher" if conn.execute("SELECT 1 FROM teachers WHERE initials = ?", (initials,)).fetchone() else "additional"
    if source == "teacher":
        excluded = conn.execute(
            "SELECT COALESCE(exclude_from_algorithm, 0) AS exclude_from_algorithm FROM teachers WHERE initials = ?",
            (initials,),
        ).fetchone()
        if excluded and excluded["exclude_from_algorithm"]:
            return "excluded from algorithm"
        if not teacher_available(conn, initials, week, day, code):
            return "not available, teaching, out of school, or already on a clashing duty"
        if duty_is_lunch(code) and teacher_lunch_limit_reached(conn, initials):
            return "maximum lunch duty limit reached"
    elif not additional_available(conn, initials, week, day, code):
        return "not active, out of school, or already on a clashing duty"
    elif role in {"ESLT", "Chaplaincy", "Admin"} and not duty_is_lunch(code):
        return "additional staff are only eligible for lunch duties"
    elif additional_max_reached(conn, initials):
        return "additional staff maximum duty limit reached"
    if source == "teacher":
        if staff_week_duty_count(conn, initials, week) >= max_duties_per_week(conn):
            return f"maximum duties per week reached ({max_duties_per_week(conn)})"
        if staff_day_duty_count(conn, initials, week, day) >= max_duties_per_day(conn):
            return f"maximum duties per day reached ({max_duties_per_day(conn)})"
    if code.startswith("Teacher_Break_Rota_") and source != "teacher":
        return "teaching staff break rota is teaching staff only"
    if code.startswith("Teacher_Break_Rota_") and teacher_break_week_count(conn, initials, week) >= 1:
        return "teacher already has one teaching staff break rota duty this week"
    if not custom_rules_allow(conn, initials, role, week, day, code):
        return "blocked by a custom hard rule"
    return None


def blank_duty_reason(conn: sqlite3.Connection, week: int, day: str, code: str) -> str:
    if code.startswith("P7_") and p7_mode(conn) == "ignore":
        return "Period 7 ignored in Rules"
    if duty_is_manual_fill_only_for_conn(conn, code):
        return "Room 90 manual fill only"
    if duty_is_optional(code):
        return "Room 90 optional"
    if code.startswith("Teacher_Break_Rota_") and int(code.rsplit("_", 1)[1]) > teacher_break_rota_slots(conn):
        return "Teaching staff break rota slot not active in settings"
    allowed_roles = role_priority_for_duty_with_settings(conn, code)
    active_roles = set()
    for row in conn.execute("SELECT initials, classification AS role FROM teachers").fetchall():
        active_roles.add(row["role"])
        if strict_assignment_allowed(conn, row["initials"], row["role"], week, day, code, "teacher"):
            return "Blank but eligible staff exist"
    for row in conn.execute(
        """
        SELECT initials, category AS role
        FROM additional_staff
        WHERE COALESCE(is_archived, 0) = 0 AND COALESCE(status, 'Active') = 'Active'
        """
    ).fetchall():
        active_roles.add(row["role"])
        if strict_assignment_allowed(conn, row["initials"], row["role"], week, day, code, "additional"):
            return "Blank but eligible staff exist"
    missing_roles = [role for role in allowed_roles if role not in active_roles]
    if missing_roles:
        return f"No active staff in required role(s): {', '.join(missing_roles)}"
    return "No eligible staff without breaking active rules"


def available_staff(conn: sqlite3.Connection, week: int, day: str, code: str) -> list[dict]:
    staff = []
    break_subjects = assigned_teacher_break_subjects(conn, week, day) if code and code.startswith("Teacher_Break_Rota_") else set()
    allowed_roles = role_priority_for_duty_with_settings(conn, code)
    balance_slt = (
        rule_active(conn, "Balanced SLT Duty Distribution")
        and allowed_roles == ["SLT"]
    )
    for row in conn.execute("SELECT initials, full_name, classification, COALESCE(subject, '') AS subject FROM teachers ORDER BY initials").fetchall():
        if strict_assignment_allowed(conn, row["initials"], row["classification"], week, day, code, "teacher"):
            staff.append({"initials": row["initials"], "name": row["full_name"], "role": row["classification"], "subject": row["subject"] or ""})
    for row in conn.execute(
        """
        SELECT initials, full_name, category FROM additional_staff
        WHERE COALESCE(is_archived, 0) = 0 AND COALESCE(status, 'Active') = 'Active'
        ORDER BY category, initials
        """
    ).fetchall():
        if strict_assignment_allowed(conn, row["initials"], row["category"], week, day, code, "additional"):
            staff.append({"initials": row["initials"], "name": row["full_name"], "role": row["category"], "subject": ""})
    if balance_slt:
        return sorted(staff, key=lambda item: slt_balance_sort_key(conn, item["initials"], week, day, code))
    if duty_uses_pastoral_distribution(code):
        return sorted(
            staff,
            key=lambda item: (
                role_priority(item["role"]),
                pastoral_distribution_sort_key(conn, item["initials"], week, day, code)
                if item["role"] == "Pastoral"
                else (999, 999, 999, 999, item["initials"]),
                item["initials"],
            ),
        )
    return sorted(
        staff,
        key=lambda item: (
            0 if break_subjects and item.get("subject", "") in break_subjects else 1,
            item.get("subject", ""),
            role_priority(item["role"]),
            item["initials"],
        ),
    )


def clear_conflicting_p4_assignments(conn: sqlite3.Connection, initials: str, week: int, day: str, code: str) -> int:
    if code.startswith("P4_Lunch_"):
        cursor = conn.execute(
            """
            UPDATE rota_assignments SET staff_initials = NULL, staff_type = NULL, assignment_source = NULL, last_updated = ?
            WHERE week = ? AND day = ? AND staff_initials = ?
              AND (period LIKE 'P4A_%' OR period LIKE 'P4B_%' OR period LIKE 'P4C_%')
            """,
            (datetime.now().isoformat(), week, day, initials),
        )
        return cursor.rowcount
    if code.startswith("P4A_") or code.startswith("P4B_") or code.startswith("P4C_"):
        cursor = conn.execute(
            """
            UPDATE rota_assignments SET staff_initials = NULL, staff_type = NULL, assignment_source = NULL, last_updated = ?
            WHERE week = ? AND day = ? AND staff_initials = ? AND period LIKE 'P4_Lunch_%'
            """,
            (datetime.now().isoformat(), week, day, initials),
        )
        return cursor.rowcount
    return 0


def assign_staff(
    conn: sqlite3.Connection,
    week: int,
    day: str,
    code: str,
    initials: str,
    role: str,
    enforce_rules: bool = True,
    assignment_source: str = "manual",
) -> int:
    conflicting_rows: list[sqlite3.Row] = []
    if code.startswith("P4_Lunch_"):
        conflicting_rows = conn.execute(
            """
            SELECT period, staff_initials, staff_type
            FROM rota_assignments
            WHERE week = ? AND day = ? AND staff_initials = ?
              AND (period LIKE 'P4A_%' OR period LIKE 'P4B_%' OR period LIKE 'P4C_%')
            """,
            (week, day, initials),
        ).fetchall()
    elif code.startswith("P4A_") or code.startswith("P4B_") or code.startswith("P4C_"):
        conflicting_rows = conn.execute(
            """
            SELECT period, staff_initials, staff_type
            FROM rota_assignments
            WHERE week = ? AND day = ? AND staff_initials = ? AND period LIKE 'P4_Lunch_%'
            """,
            (week, day, initials),
        ).fetchall()
    cleared = 0
    if conflicting_rows:
        cleared = clear_conflicting_p4_assignments(conn, initials, week, day, code)
    if enforce_rules and not strict_assignment_allowed(conn, initials, role, week, day, code):
        for row in conflicting_rows:
            conn.execute(
                """
                UPDATE rota_assignments
                SET staff_initials = ?, staff_type = ?, assignment_source = COALESCE(assignment_source, 'manual'), last_updated = ?
                WHERE week = ? AND day = ? AND period = ?
                """,
                (row["staff_initials"], row["staff_type"], datetime.now().isoformat(), week, day, row["period"]),
            )
        conn.commit()
        return -1
    conn.execute(
        """
        INSERT INTO rota_assignments(week, day, period, staff_initials, staff_type, assignment_source, last_updated)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(week, day, period)
        DO UPDATE SET
            staff_initials = excluded.staff_initials,
            staff_type = excluded.staff_type,
            assignment_source = excluded.assignment_source,
            last_updated = excluded.last_updated
        """,
        (week, day, code, initials, role, assignment_source, datetime.now().isoformat()),
    )
    conn.commit()
    return cleared


def get_p4_lunch_conflicts(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT week, day, staff_initials, GROUP_CONCAT(period, ', ') AS duties
        FROM rota_assignments
        WHERE staff_initials IS NOT NULL
          AND (period LIKE 'P4_Lunch_%' OR period LIKE 'P4A_%' OR period LIKE 'P4B_%' OR period LIKE 'P4C_%')
        GROUP BY week, day, staff_initials
        HAVING SUM(CASE WHEN period LIKE 'P4_Lunch_%' THEN 1 ELSE 0 END) > 0
           AND SUM(CASE WHEN period LIKE 'P4A_%' OR period LIKE 'P4B_%' OR period LIKE 'P4C_%' THEN 1 ELSE 0 END) > 0
        ORDER BY week, day, staff_initials
        """
    ).fetchall()


def repair_p4_lunch_conflicts(conn: sqlite3.Connection) -> int:
    conflicts = get_p4_lunch_conflicts(conn)
    cleared = 0
    for row in conflicts:
        cursor = conn.execute(
            """
            UPDATE rota_assignments
            SET staff_initials = NULL, staff_type = NULL, assignment_source = NULL, last_updated = ?
            WHERE week = ? AND day = ? AND staff_initials = ?
              AND (period LIKE 'P4A_%' OR period LIKE 'P4B_%' OR period LIKE 'P4C_%')
            """,
            (datetime.now().isoformat(), row["week"], row["day"], row["staff_initials"]),
        )
        cleared += cursor.rowcount
    conn.commit()
    return cleared


def duty_sort_key(slot: sqlite3.Row | tuple[int, str, str]) -> tuple:
    week, day, code = (slot["week"], slot["day"], slot["period"]) if hasattr(slot, "keys") else slot
    priority = 50
    if code == "Gate":
        priority = 0
    elif code == "P1_Isolation":
        priority = 1
    elif code == "Tutor_1st_Duty":
        priority = 2
    elif code == "Break_Duty_Lead":
        priority = 3
    elif code.startswith("Teacher_Break_Rota_"):
        priority = 4
    elif code.startswith("P7_Detention"):
        priority = 5
    elif code.startswith("P4_Lunch"):
        priority = 6
    elif "Isolation" in code:
        priority = 7
    elif duty_is_optional_for_conn(conn, code):
        priority = 99
    return (priority, week, ROTA_DAYS.index(day), DUTY_ORDER.get(code, 999))


def eligible_teaching_staff_count(conn: sqlite3.Connection, week: int, day: str, code: str) -> int:
    count = 0
    for row in conn.execute("SELECT initials, classification AS role FROM teachers").fetchall():
        if strict_assignment_allowed(conn, row["initials"], row["role"], week, day, code, "teacher"):
            count += 1
    return count


def revised_slot_sort_key(conn: sqlite3.Connection, slot: sqlite3.Row) -> tuple:
    code = slot["period"]
    if duty_is_lunch(code):
        return (0, slot["week"], ROTA_DAYS.index(slot["day"]), DUTY_ORDER.get(code, 999))
    return (
        1,
        eligible_teaching_staff_count(conn, slot["week"], slot["day"], code),
        slot["id"] if "id" in slot.keys() else DUTY_ORDER.get(code, 999),
    )


def choose_lunch_candidate(conn: sqlite3.Connection, eligible: list[dict], code: str, week: int, day: str) -> dict | None:
    if code == "P4_Lunch_1":
        eligible.sort(
            key=lambda cand: (
                staff_total_duty_count(conn, cand["initials"]),
                previous_non_free_streak(conn, cand["initials"], week, day, code),
                cand["initials"],
            )
        )
        return eligible[0] if eligible else None
    additional_priority = {"ESLT": 0, "Chaplaincy": 1, "Admin": 2}
    additional = [cand for cand in eligible if cand["role"] in additional_priority]
    if additional:
        additional.sort(
            key=lambda cand: (
                additional_priority[cand["role"]],
                0 if staff_total_duty_count(conn, cand["initials"]) < int(cand.get("min_duties") or 0) else 1,
                staff_total_duty_count(conn, cand["initials"]),
                previous_non_free_streak(conn, cand["initials"], week, day, code),
                cand["initials"],
            )
        )
    teaching = [cand for cand in eligible if cand["source"] == "teacher"]
    if code != "P4_Lunch_1" and additional:
        return additional[0]
    if teaching:
        scored = []
        for cand in teaching:
            score, tie_break = teacher_assignment_score(conn, cand["initials"], week, day, code)
            scored.append((-score, -tie_break, cand["initials"], cand))
        scored.sort()
        return scored[0][-1]
    return additional[0] if additional else None


def auto_assign_empty_slots(conn: sqlite3.Connection) -> dict:
    ensure_duty_event_rows(conn)
    pre_repaired = repair_p4_lunch_conflicts(conn)

    candidates = []
    for row in conn.execute(
        """
        SELECT initials, classification AS role, total_lessons, protected_periods, days_in_school, COALESCE(subject, '') AS subject
        FROM teachers
        """
    ).fetchall():
        days = (row["days_in_school"] or "1111111111").ljust(10, "1")[:10]
        max_load = FULL_TIME_TEACHING_LOAD - (DAILY_TEACHING_LOAD * days.count("0"))
        current_duties = conn.execute("SELECT COUNT(*) AS count FROM rota_assignments WHERE staff_initials = ?", (row["initials"],)).fetchone()["count"]
        heavy_rows = conn.execute("SELECT period FROM rota_assignments WHERE staff_initials = ?", (row["initials"],)).fetchall()
        heavy_count = sum(1 for heavy in heavy_rows if duty_is_heavy(heavy["period"]))
        candidates.append(
            {
                "initials": row["initials"],
                "role": row["role"],
                "available": max_load - float(row["total_lessons"] or 0) - int(row["protected_periods"] or 0) - current_duties,
                "duties": current_duties,
                "heavy": heavy_count,
                "source": "teacher",
                "subject": row["subject"] or "",
            }
        )

    for row in conn.execute(
        """
        SELECT initials, category AS role, min_duties, max_duties FROM additional_staff
        WHERE COALESCE(is_archived, 0) = 0 AND COALESCE(status, 'Active') = 'Active'
        """
    ).fetchall():
        current_duties = conn.execute("SELECT COUNT(*) AS count FROM rota_assignments WHERE staff_initials = ?", (row["initials"],)).fetchone()["count"]
        heavy_rows = conn.execute("SELECT period FROM rota_assignments WHERE staff_initials = ?", (row["initials"],)).fetchall()
        heavy_count = sum(1 for heavy in heavy_rows if duty_is_heavy(heavy["period"]))
        candidates.append(
            {
                "initials": row["initials"],
                "role": row["role"],
                "available": 999,
                "duties": current_duties,
                "heavy": heavy_count,
                "source": "additional",
                "min_duties": int(row["min_duties"] or 0),
                "max_duties": row["max_duties"],
            }
        )

    empty_slots = conn.execute(
        "SELECT id, week, day, period FROM rota_assignments WHERE staff_initials IS NULL"
    ).fetchall()
    active_codes = active_duty_codes(conn)
    assigned = 0
    issues = []
    for slot in sorted(empty_slots, key=lambda item: revised_slot_sort_key(conn, item)):
        week, day, code = slot["week"], slot["day"], slot["period"]
        if code not in active_codes:
            continue
        if duty_is_manual_fill_only_for_conn(conn, code):
            continue
        allowed_roles = role_priority_for_duty_with_settings(conn, code)
        balance_slt = (
            rule_active(conn, "Balanced SLT Duty Distribution")
            and allowed_roles == ["SLT"]
        )
        eligible = []
        break_subjects = assigned_teacher_break_subjects(conn, week, day) if code.startswith("Teacher_Break_Rota_") else set()
        for cand in candidates:
            if not strict_assignment_allowed(conn, cand["initials"], cand["role"], week, day, code, cand["source"]):
                continue
            if duty_is_lunch(code) and cand["source"] == "additional" and cand["role"] not in {"ESLT", "Chaplaincy", "Admin"}:
                continue
            if not duty_is_lunch(code) and cand["source"] == "additional" and cand["role"] in {"ESLT", "Chaplaincy", "Admin"}:
                continue
            if cand["source"] == "teacher":
                score, tie_break = teacher_assignment_score(conn, cand["initials"], week, day, code)
                subject_score = 0
                if code.startswith("Teacher_Break_Rota_") and break_subjects:
                    subject_score = 0 if cand.get("subject", "") in break_subjects else 1
                if balance_slt:
                    sort_key = slt_balance_sort_key(conn, cand["initials"], week, day, code)
                elif duty_uses_pastoral_distribution(code):
                    pastoral_key = (
                        pastoral_distribution_sort_key(conn, cand["initials"], week, day, code)
                        if cand["role"] == "Pastoral"
                        else (999, 999, 999, 999, cand["initials"])
                    )
                    sort_key = (
                        role_priority(cand["role"]),
                        pastoral_key[0],
                        pastoral_key[1],
                        pastoral_key[2],
                        pastoral_key[3],
                        -score,
                        -tie_break,
                        cand["initials"],
                    )
                else:
                    sort_key = (subject_score, role_priority(cand["role"]), -score, -tie_break, cand["initials"])
                eligible.append({"sort": sort_key, **cand})
            else:
                if balance_slt:
                    eligible.append({"sort": slt_balance_sort_key(conn, cand["initials"], week, day, code), **cand})
                elif duty_uses_pastoral_distribution(code):
                    pastoral_key = (
                        pastoral_distribution_sort_key(conn, cand["initials"], week, day, code)
                        if cand["role"] == "Pastoral"
                        else (999, 999, 999, 999, cand["initials"])
                    )
                    eligible.append({
                        "sort": (
                            role_priority(cand["role"]),
                            pastoral_key[0],
                            pastoral_key[1],
                            pastoral_key[2],
                            pastoral_key[3],
                            staff_total_duty_count(conn, cand["initials"]),
                            cand["initials"],
                        ),
                        **cand,
                    })
                else:
                    eligible.append({"sort": (role_priority(cand["role"]), staff_total_duty_count(conn, cand["initials"]), cand["initials"]), **cand})
        if not eligible:
            if not duty_is_optional_for_conn(conn, code):
                issues.append(slot)
            continue
        if duty_is_lunch(code):
            chosen = choose_lunch_candidate(conn, eligible, code, week, day)
        else:
            eligible.sort(key=lambda cand: cand.get("sort", (999, cand["initials"])))
            chosen = eligible[0]
        if not chosen:
            if not duty_is_optional_for_conn(conn, code):
                issues.append(slot)
            continue
        conn.execute(
            """
            UPDATE rota_assignments
            SET staff_initials = ?, staff_type = ?, assignment_source = 'auto', last_updated = ?
            WHERE week = ? AND day = ? AND period = ?
            """,
            (chosen["initials"], chosen["role"], datetime.now().isoformat(), week, day, code),
        )
        chosen["duties"] += 1
        chosen["available"] -= 1
        if duty_is_heavy(code):
            chosen["heavy"] += 1
        assigned += 1

    for slot in issues:
        conn.execute(
            """
            INSERT INTO problem_log(issue_type, description, week, day, period)
            VALUES (?, ?, ?, ?, ?)
            """,
            ("Insufficient Staff", f"Auto-assign could not fill {DUTY_LABELS.get(slot['period'], slot['period'])}", slot["week"], slot["day"], slot["period"]),
        )
    conn.commit()
    post_repaired = repair_p4_lunch_conflicts(conn)
    return {"assigned": assigned, "issues": len(issues), "repaired": pre_repaired + post_repaired}


def preview_auto_assign(conn: sqlite3.Connection) -> dict:
    memory = sqlite3.connect(":memory:")
    memory.row_factory = sqlite3.Row
    conn.backup(memory)
    before = {
        (row["week"], row["day"], row["period"]): row["staff_initials"]
        for row in memory.execute("SELECT week, day, period, staff_initials FROM rota_assignments").fetchall()
    }
    result = auto_assign_empty_slots(memory)
    active_codes = active_duty_codes(memory)
    after_rows = memory.execute("SELECT week, day, period, staff_initials, staff_type FROM rota_assignments").fetchall()
    would_assign = []
    blanks = []
    for row in after_rows:
        key = (row["week"], row["day"], row["period"])
        if row["period"] not in active_codes:
            continue
        if not before.get(key) and row["staff_initials"]:
            would_assign.append(dict(row))
        elif not row["staff_initials"]:
            blanks.append(
                {
                    "week": row["week"],
                    "day": row["day"],
                    "period": row["period"],
                    "reason": blank_duty_reason(memory, row["week"], row["day"], row["period"]),
                }
            )
    memory.close()
    return {"result": result, "would_assign": would_assign, "blanks": blanks}


def workload_summary(conn: sqlite3.Connection) -> list[dict]:
    staff: dict[str, dict] = {}
    for row in conn.execute(
        """
        SELECT initials, COALESCE(full_name, initials) AS name, classification AS role,
               total_lessons, protected_periods, days_in_school
        FROM teachers
        """
    ).fetchall():
        days = (row["days_in_school"] or "1111111111").ljust(10, "1")[:10]
        max_load = FULL_TIME_TEACHING_LOAD - (DAILY_TEACHING_LOAD * days.count("0"))
        staff[row["initials"]] = {
            "initials": row["initials"],
            "name": row["name"],
            "role": row["role"],
            "available_periods": max_load - float(row["total_lessons"] or 0) - int(row["protected_periods"] or 0),
            "duties": 0,
            "heavy": 0,
            "lunch": 0,
            "isolation": 0,
        }
    for row in conn.execute(
        """
        SELECT initials, full_name AS name, category AS role
        FROM additional_staff
        WHERE COALESCE(is_archived, 0) = 0
        """
    ).fetchall():
        staff[row["initials"]] = {
            "initials": row["initials"],
            "name": row["name"],
            "role": row["role"],
            "available_periods": 999,
            "duties": 0,
            "heavy": 0,
            "lunch": 0,
            "isolation": 0,
        }
    for row in conn.execute("SELECT staff_initials, period FROM rota_assignments WHERE staff_initials IS NOT NULL").fetchall():
        item = staff.get(row["staff_initials"])
        if not item:
            continue
        item["duties"] += 1
        if duty_is_heavy(row["period"]):
            item["heavy"] += 1
        if "Lunch" in row["period"]:
            item["lunch"] += 1
        if "Isolation" in row["period"]:
            item["isolation"] += 1
    rows = list(staff.values())
    role_groups: dict[str, list[int]] = {}
    for row in rows:
        role_groups.setdefault(row["role"], []).append(row["duties"])
    role_avgs = {role: (sum(values) / len(values) if values else 0) for role, values in role_groups.items()}
    for row in rows:
        diff = row["duties"] - role_avgs.get(row["role"], 0)
        row["outlier"] = "High" if diff > 3 else ("Low" if diff < -3 else "")
    return sorted(rows, key=lambda item: (-item["duties"], -item["heavy"], item["initials"]))


def fairness_summary(workload: list[dict]) -> dict:
    active = [row for row in workload if row["duties"] > 0]
    if not active:
        return {"label": "Not built yet", "detail": "No duties have been assigned."}
    duties = [row["duties"] for row in active]
    spread = max(duties) - min(duties)
    if spread <= 2:
        label = "Good"
    elif spread <= 5:
        label = "Needs review"
    else:
        label = "Uneven"
    return {"label": label, "detail": f"Duty spread is {spread} between the least and most-used assigned staff."}


def proposed_review(conn: sqlite3.Connection) -> dict:
    ensure_duty_event_rows(conn)
    active_codes = active_duty_codes(conn)
    blanks = []
    for row in conn.execute(
        "SELECT week, day, period FROM rota_assignments WHERE staff_initials IS NULL ORDER BY week, day, period"
    ).fetchall():
        if row["period"] not in active_codes:
            continue
        blanks.append(
            {
                "week": row["week"],
                "day": row["day"],
                "period": row["period"],
                "reason": blank_duty_reason(conn, row["week"], row["day"], row["period"]),
            }
        )
    workload = workload_summary(conn)
    lunch_additional = conn.execute(
        """
        SELECT r.week, r.day, r.period, r.staff_initials, r.staff_type
        FROM rota_assignments r
        WHERE r.period LIKE 'P4_Lunch_%'
          AND r.staff_type IN ('ESLT', 'Chaplaincy', 'Admin')
        ORDER BY r.week, r.day, r.period
        """
    ).fetchall()
    return {
        "blanks": blanks,
        "conflicts": get_p4_lunch_conflicts(conn),
        "workload": workload,
        "fairness": fairness_summary(workload),
        "lunch_additional": lunch_additional,
    }


def build_master_style_workbook(conn: sqlite3.Connection) -> BytesIO:
    ensure_duty_event_rows(conn)
    assignment_rows = conn.execute("SELECT week, day, period, staff_initials FROM rota_assignments").fetchall()
    assignments = {(row["week"], row["day"], row["period"]): row["staff_initials"] or "" for row in assignment_rows}

    template_path = MASTER_TEMPLATE_PATH if MASTER_TEMPLATE_PATH.exists() else Path("2025 MASTER DUTY FINAL.xlsx")
    if template_path.exists():
        wb = load_workbook(template_path)
        ws = wb["Master"]
        day_columns = {
            (1, "Mon"): 2, (1, "Tue"): 3, (1, "Wed"): 4, (1, "Thu"): 5, (1, "Fri"): 6,
            (2, "Mon"): 7, (2, "Tue"): 8, (2, "Wed"): 9, (2, "Thu"): 10, (2, "Fri"): 11,
        }
        staff_rows = [3, 5, 7, 8, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29, 31, 34, 35, 37, 39,
                      41, 43, 45, 47, 49, 51, 53, 55, 57, 59, 60, 61, 62, 63, 64, 65, 67, 71,
                      73, 75, 77, 80, 82, 84, 86, 88, 90, 92, 94, 96, 98, 102]
        for row in staff_rows:
            for col in range(2, 12):
                ws.cell(row, col).value = None
        for row in range(12, ws.max_row + 1):
            for col in range(17, 24):
                ws.cell(row, col).value = None
        master_row_map = {
            "Tutor_1st_Duty": 3, "Tutor_AOW": 7, "Tutor_Pastoral_Support": 11, "Tutor_Room_90": 13,
            "P1_First_Duty": 15, "P1_Pastoral_Support": 17, "P1_Room_90": 19, "P1_Isolation": 21,
            "P2_First_Duty": 23, "P2_Pastoral_Support": 25, "P2_Room_90": 27, "P2_Isolation": 29,
            "Break_Duty_Lead": 31, "Break_Late_Detention": 34, "Break_Pastoral_Support": 37, "Break_Room_90": 39,
            "P3_First_Duty": 41, "P3_Pastoral_Support": 43, "P3_Room_90": 45, "P3_Isolation": 47,
            "P4A_First_Duty": 49, "P4A_Pastoral_Support": 51, "P4A_Isolation": 55,
            "P4_Lunch_1": 57, "P4_Lunch_2": 59, "P4_Lunch_3": 60, "P4_Lunch_4": 61,
            "P4_Lunch_5": 62, "P4_Lunch_6": 63, "P4_Lunch_7": 64,
            "P4B_First_Duty": 65, "P4B_Pastoral_Support": 67, "P4B_Isolation": 71,
            "P4C_First_Duty": 73, "P4C_Pastoral_Support": 75, "P4C_Isolation": 77, "P4C_Rest_Break": 80,
            "P5_First_Duty": 82, "P5_Pastoral_Support": 84, "P5_Room_90": 86, "P5_Isolation": 88,
            "P6_First_Duty": 90, "P6_Pastoral_Support": 92, "P6_Room_90": 94, "P6_Isolation": 96,
            "P7_Homework_Club": 98,
        }
        for code, row in master_row_map.items():
            for (week, day), col in day_columns.items():
                ws.cell(row, col).value = assignments.get((week, day, code), "")
        if "DutyMaster extra duties" in wb.sheetnames:
            del wb["DutyMaster extra duties"]
        extra = wb.create_sheet("DutyMaster extra duties")
        extra.append(["Week", "Day", "Duty", "Staff"])
        extra_codes = [
            "Gate", "Tutor_Isolation", "Break_Isolation",
            *[f"Teacher_Break_Rota_{index}" for index in range(1, teacher_break_rota_slots(conn) + 1)],
            "P4A_Rest_Break", "P4B_Rest_Break", "P7_Detention_1", "P7_Detention_2",
        ]
        for code in extra_codes:
            for week in ROTA_WEEKS:
                for day in ROTA_DAYS:
                    extra.append([week, day, DUTY_LABELS.get(code, code), assignments.get((week, day, code), "")])
        for col in range(1, 5):
            extra.column_dimensions[get_column_letter(col)].width = 18
            extra.cell(1, col).font = Font(bold=True)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    wb = Workbook()
    ws = wb.active
    ws.title = "Master"
    fills = {
        "header": PatternFill("solid", fgColor="1F4E78"),
        "section": PatternFill("solid", fgColor="E2F0D9"),
    }
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.append(["Duty / Event", "Week 1 Mon", "Week 1 Tue", "Week 1 Wed", "Week 1 Thu", "Week 1 Fri",
               "Week 2 Mon", "Week 2 Tue", "Week 2 Wed", "Week 2 Thu", "Week 2 Fri"])
    for cell in ws[1]:
        cell.fill = fills["header"]
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for section, events in active_duty_sections(conn):
        ws.append([section])
        ws.cell(ws.max_row, 1).fill = fills["section"]
        for code, label in events:
            row = [label]
            for week in ROTA_WEEKS:
                for day in ROTA_DAYS:
                    row.append(assignments.get((week, day, code), ""))
            ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for idx in range(1, 12):
        ws.column_dimensions[get_column_letter(idx)].width = 16
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
